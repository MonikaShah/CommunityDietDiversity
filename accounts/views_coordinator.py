import io
import openpyxl
import xlsxwriter
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.models import Group, User
from django.contrib.auth.decorators import login_required, user_passes_test
from .decorators import *
from .models import *
from .forms import *
from .helper_functions import *
from shared.encryption import EncryptionHelper
from django.utils import timezone

encryptionHelper = EncryptionHelper()


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_coordinator, login_url="accounts:forbidden")
@password_change_required
def coordinator_dashboard(request):
    teachers = (
        CoordinatorInCharge.objects.filter(user=request.user)
        .first()
        .teacherincharge_set.all()
    )
    for teacher in teachers:
        teacher.name = encryptionHelper.decrypt(teacher.name)
    return render(
        request,
        "coordinator/coordinator_dashboard.html",
        {"teachers": teachers, "page_type": "coordinator_dashboard"},
    )


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_coordinator, login_url="accounts:forbidden")
@password_change_required
def addTeacherForm(request):
    if request.method == "GET":
        form = TeachersInfoForm()
        user_creation_form = UserCreationForm()
        return render(
            request,
            "coordinator/add_teacher.html",
            {"form": form, "user_creation_form": user_creation_form},
        )
    else:
        form = TeachersInfoForm(request.POST)
        teacheruserform = UserCreationForm(request.POST)
        if form.is_valid() and teacheruserform.is_valid():
            teacheruser = teacheruserform.save(commit=False)
            teacheruser.save()
            teacher_group = Group.objects.get(name="Teachers")
            teacheruser.groups.add(teacher_group)
            teacheruser.save()
            teacher = form.save(commit=False)
            teacher.user = teacheruser
            teacher.email = encryptionHelper.encrypt(request.POST["email"])
            teacher.name = encryptionHelper.encrypt(request.POST["name"])
            teacher.dob = encryptionHelper.encrypt(request.POST["dob"])
            teacher.gender = encryptionHelper.encrypt(request.POST["gender"])
            teacher.mobile_no = encryptionHelper.encrypt(request.POST["mobile_no"])
            teacher.organization = Organization.objects.filter(
                name=CoordinatorInCharge.objects.filter(user=request.user)
                .first()
                .organization
            ).first()
            teacher.coordinator = CoordinatorInCharge.objects.filter(
                user=request.user
            ).first()
            teacher.save()
            return redirect("accounts:coordinator_dashboard")
        else:
            return render(
                request,
                "coordinator/add_teacher.html",
                {"form": form, "user_creation_form": teacheruserform},
            )


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_coordinator, login_url="accounts:forbidden")
@password_change_required
def allSessions(request):
    sessions = (
        CoordinatorInCharge.objects.filter(user=request.user).first().session_set.all()
    )
    open_sessions = []
    close_sessions = []
    upcoming_sessions = []
    open = False
    close = False
    upcoming = False
    for session in sessions:
        if session.start_date > timezone.now():
            upcoming_sessions.append(session)
            upcoming = True
        elif session.end_date == None:
            open_sessions.append(session)
            open = True
        else:
            close_sessions.append(session)
            close = True
    return render(
        request,
        "coordinator/all_sessions.html",
        {
            "open_sessions": open_sessions,
            "upcoming_sessions": upcoming_sessions,
            "close_sessions": close_sessions,
            "open": open,
            "upcoming": upcoming,
            "close": close,
            "page_type": "all_sessions",
        },
    )


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_coordinator, login_url="accounts:forbidden")
@password_change_required
def addSessionForm(request):
    if request.method == "GET":
        form = SessionsInfoForm()
        return render(
            request,
            "coordinator/add_session.html",
            {"form": form},
        )
    else:
        form = SessionsInfoForm(request.POST)
        if form.is_valid():
            session = form.save(commit=False)
            session.coordinator = CoordinatorInCharge.objects.filter(
                user=request.user
            ).first()
            session.save()
            return redirect("accounts:all_sessions")
        else:
            return render(
                request,
                "coordinator/add_session.html",
                {"form": form},
            )


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_coordinator, login_url="accounts:forbidden")
@password_change_required
def viewSessionTeachers(request, id, open_id, my_messages=None):
    session = Session.objects.filter(id=id).first()
    objects = Teacher_Session.objects.filter(session=session)
    teachers = []
    for object in objects:
        object.teacher.name = encryptionHelper.decrypt(object.teacher.name)
        teachers.append(object.teacher)
    if my_messages != None:
        return render(
            request,
            "coordinator/view_session_teachers.html",
            {
                "teachers": teachers,
                "session": session,
                "my_messages": my_messages,
                "open_id": open_id,
                "page_type": "view_session_teachers",
            },
        )
    else:
        return render(
            request,
            "coordinator/view_session_teachers.html",
            {
                "teachers": teachers,
                "session": session,
                "open_id": open_id,
                "page_type": "view_session_teachers",
            },
        )


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_coordinator, login_url="accounts:forbidden")
@password_change_required
def getSessionTeachersTemplate(request):
    output = io.BytesIO()
    wb = xlsxwriter.Workbook(output)
    ws = wb.add_worksheet("Session Teachers Data")
    columns = [
        "Teacher Username",
    ]
    for col_num in range(len(columns)):
        ws.write(0, col_num, columns[col_num])
    wb.close()
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = "attachment; filename=sessionTeacherTemplate.xlsx"
    return response


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_coordinator, login_url="accounts:forbidden")
@password_change_required
def addSessionTeachers(request, id):
    if request.method == "GET":
        return render(
            request,
            "coordinator/add_session_teachers.html",
            {"page_type": "add_session_teachers"},
        )
    else:
        try:
            excel_file = request.FILES["excel_file"]
            if excel_file.name[-4:] == "xlsx":
                try:
                    wb = openpyxl.load_workbook(excel_file)
                    teacherSheet = wb["Session Teachers Data"]
                except:
                    return render(
                        request,
                        "coordinator/add_session_teachers.html",
                        {
                            "page_type": "add_session_teachers",
                            "my_messages": {
                                "error": "Incorrect file uploaded, please use the template provided above."
                            },
                        },
                    )
            else:
                return render(
                    request,
                    "coordinator/add_session_teachers.html",
                    {
                        "page_type": "add_session_teachers",
                        "my_messages": {
                            "error": "Incorrect file type, only .xlsx files are allowed!"
                        },
                    },
                )
        except:
            return render(
                request,
                "coordinator/add_session_teachers.html",
                {
                    "page_type": "add_session_teachers",
                    "my_messages": {"error": "Sorry something went wrong!"},
                },
            )

        organization = (
            CoordinatorInCharge.objects.filter(user=request.user).first().organization
        )
        breaking = error = False
        error_message = ""
        teacher_data = []
        for row_no, row in enumerate(teacherSheet.iter_rows()):
            if breaking == True:
                break
            if row_no == 0:
                continue
            for cell in row:
                if cell.column_letter == "A":
                    if cell.value == None:
                        breaking = True
                        break
                    else:
                        if User.objects.filter(username=cell.value).exists():
                            user = User.objects.filter(username=cell.value).first()
                            if is_teacher(user):
                                teacher_user = TeacherInCharge.objects.filter(
                                    user=user
                                ).first()
                                if teacher_user.organization == organization:
                                    teacher_data.append(teacher_user)
                                else:
                                    breaking = error = True
                                    error_message = (
                                        "Teacher at row number "
                                        + str(row_no + 1)
                                        + " does not belong to your organization"
                                    )
                                    break
                            else:
                                breaking = error = True
                                error_message = (
                                    "User at row number "
                                    + str(row_no + 1)
                                    + " is not a teacher"
                                )
                                break
                        else:
                            breaking = error = True
                            error_message = "Invalid username at row number " + str(
                                row_no + 1
                            )
                            break

        if error == True:
            return render(
                request,
                "coordinator/add_session_teachers.html",
                {
                    "page_type": "add_session_teachers",
                    "my_messages": {"error": error_message},
                },
            )

        session = Session.objects.filter(id=id).first()
        teacher_session = Teacher_Session.objects.filter(session=session)
        teacher_session_teacher_objects = []
        for i in teacher_session:
            teacher_session_teacher_objects.append(i.teacher)
        teacher_added = 0
        teacher_exists = 0
        for teacher_user in teacher_data:
            if teacher_user in teacher_session_teacher_objects:
                teacher_exists += 1
            else:
                new_teacher_session = Teacher_Session()
                new_teacher_session.session = session
                new_teacher_session.teacher = teacher_user
                new_teacher_session.save()
                teacher_added += 1

        if teacher_added == 0:
            my_messages = {
                "success": "Registration successful. Already registered: "
                + str(teacher_exists)
            }
        elif teacher_exists == 0:
            my_messages = {
                "success": "Registration successful. Newly registered: "
                + str(teacher_added)
            }
        else:
            my_messages = {
                "success": "Registration successful. Already registered: "
                + str(teacher_exists)
                + ", Newly registered: "
                + str(teacher_added)
            }

        return viewSessionTeachers(request, id, 1, my_messages)


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_coordinator, login_url="accounts:forbidden")
@password_change_required
def addSessionTeachersList(request, id):
    if request.method == "GET":
        teachers_in_sessions = Teacher_Session.objects.all()
        teachers_in_sessions_id = []
        coordinator = CoordinatorInCharge.objects.filter(user=request.user).first()
        organization = coordinator.organization
        for teacher in teachers_in_sessions:
            teachers_in_sessions_id.append(teacher.teacher.id)

        teachers = TeacherInCharge.objects.filter(
            coordinator=coordinator, organization=organization
        ).exclude(id__in=teachers_in_sessions_id)

        for teacher in teachers:
            teacher.name = encryptionHelper.decrypt(teacher.name)
        return render(
            request,
            "coordinator/add_session_teachers_list.html",
            {"page_type": "add_session_teachers_list", "teachers": teachers},
        )
    else:
        teachers_id_list = request.POST.getlist("chk")
        session = Session.objects.filter(id=id).first()
        for t_id in teachers_id_list:
            teacher = TeacherInCharge.objects.filter(id=t_id).first()
            teacher_session = Teacher_Session()
            teacher_session.session = session
            teacher_session.teacher = teacher
            teacher_session.save()

        return redirect("accounts:view_session_teachers", id, 1)


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_coordinator, login_url="accounts:forbidden")
@password_change_required
def removeSessionTeacher(request, session_id, teacher_id):
    teacher = TeacherInCharge.objects.filter(id=teacher_id).first()
    Teacher_Session.objects.filter(teacher=teacher).delete()
    my_messages = {"success": "Teacher removed from session successfully"}
    return viewSessionTeachers(request, session_id, 1, my_messages)


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_coordinator, login_url="accounts:forbidden")
@password_change_required
def coordinator_reset_password(request):
    if request.method == "GET":
        form = CoordPasswordReset()
        return render(
            request,
            "coordinator/coordinator_reset_password.html",
            {"form": form, "page_type": "reset_password"},
        )
    else:
        form = CoordPasswordReset(request.POST)
        organization = CoordinatorInCharge.objects.get(user=request.user).organization
        if form.is_valid():
            username = form.cleaned_data["username"]
            if User.objects.filter(username=username).exists():
                user = User.objects.get(username=username)
                if is_teacher(user):
                    user_teacher = TeacherInCharge.objects.get(user=user)
                    if user_teacher.organization == organization:
                        password = random_password_generator()
                        user_teacher.first_password = password
                        user_teacher.password_changed = False
                        user_teacher.save()
                        user.set_password(password)
                        user.save()
                        return render(
                            request,
                            "coordinator/coordinator_reset_password.html",
                            {
                                "form": form,
                                "page_type": "reset_password",
                                "my_messages": {
                                    "success": "Password reset successfull. Download login credentions from below."
                                },
                            },
                        )
                    else:
                        return render(
                            request,
                            "coordinator/coordinator_reset_password.html",
                            {
                                "form": form,
                                "page_type": "reset_password",
                                "my_messages": {
                                    "error": "Teacher does not belong to your organization."
                                },
                            },
                        )
                elif is_parent(user):
                    user_parent = ParentsInfo.objects.get(user=user)
                    if user_parent.organization == organization:
                        password = random_password_generator()
                        user_parent.first_password = password
                        user_parent.password_changed = False
                        user_parent.save()
                        user.set_password(password)
                        user.save()
                        return render(
                            request,
                            "coordinator/coordinator_reset_password.html",
                            {
                                "form": form,
                                "page_type": "reset_password",
                                "my_messages": {
                                    "success": "Password reset successfull. Download login credentions from below."
                                },
                            },
                        )
                    else:
                        return render(
                            request,
                            "coordinator/coordinator_reset_password.html",
                            {
                                "form": form,
                                "page_type": "reset_password",
                                "my_messages": {
                                    "error": "Parent does not belong to your organization."
                                },
                            },
                        )
                elif is_student(user):
                    user_student = StudentsInfo.objects.get(user=user)
                    if user_student.organization == organization:
                        password = random_password_generator()
                        user_student.first_password = password
                        user_student.password_changed = False
                        user_student.save()
                        user.set_password(password)
                        user.save()
                        return render(
                            request,
                            "coordinator/coordinator_reset_password.html",
                            {
                                "form": form,
                                "page_type": "reset_password",
                                "my_messages": {
                                    "success": "Password reset successfull. Download login credentions from below."
                                },
                            },
                        )
                    else:
                        return render(
                            request,
                            "coordinator/coordinator_reset_password.html",
                            {
                                "form": form,
                                "page_type": "reset_password",
                                "my_messages": {
                                    "error": "Student does not belong to your organization."
                                },
                            },
                        )
                else:
                    return render(
                        request,
                        "coordinator/coordinator_reset_password.html",
                        {
                            "form": form,
                            "page_type": "reset_password",
                            "my_messages": {
                                "error": "Provided user is not a teacher/parent/student."
                            },
                        },
                    )
            else:
                return render(
                    request,
                    "coordinator/coordinator_reset_password.html",
                    {
                        "form": form,
                        "page_type": "reset_password",
                        "my_messages": {"error": "Invalid username."},
                    },
                )
        else:
            return render(
                request,
                "coordinator/coordinator_reset_password.html",
                {"form": form, "page_type": "reset_password"},
            )


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_coordinator, login_url="accounts:forbidden")
@password_change_required
def coordinator_reset_password_teacher_download(request):
    output = io.BytesIO()
    wb = xlsxwriter.Workbook(output)
    bold = wb.add_format({"bold": True})
    credentials = wb.add_worksheet("Credentials")
    columns = [
        "Username",
        "Password",
    ]
    for col_num in range(len(columns)):
        credentials.write(0, col_num, columns[col_num], bold)
    teacher = TeacherInCharge.objects.filter(password_changed=False)
    for row_no, x in enumerate(teacher):
        credentials.write(row_no + 1, 0, x.user.username)
        credentials.write(row_no + 1, 1, x.first_password)
    wb.close()
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response[
        "Content-Disposition"
    ] = "attachment; filename=Teacher Login Credentials.xlsx"
    return response


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_coordinator, login_url="accounts:forbidden")
@password_change_required
def coordinator_reset_password_parent_download(request):
    output = io.BytesIO()
    wb = xlsxwriter.Workbook(output)
    bold = wb.add_format({"bold": True})
    credentials = wb.add_worksheet("Credentials")
    columns = [
        "Username",
        "Password",
    ]
    for col_num in range(len(columns)):
        credentials.write(0, col_num, columns[col_num], bold)
    parent = ParentsInfo.objects.filter(password_changed=False)
    for row_no, x in enumerate(parent):
        credentials.write(row_no + 1, 0, x.user.username)
        credentials.write(row_no + 1, 1, x.first_password)
    wb.close()
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response[
        "Content-Disposition"
    ] = "attachment; filename=Parent Login Credentials.xlsx"
    return response


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_coordinator, login_url="accounts:forbidden")
@password_change_required
def coordinator_reset_password_student_download(request):
    output = io.BytesIO()
    wb = xlsxwriter.Workbook(output)
    bold = wb.add_format({"bold": True})
    credentials = wb.add_worksheet("Credentials")
    columns = [
        "Username",
        "Password",
    ]
    for col_num in range(len(columns)):
        credentials.write(0, col_num, columns[col_num], bold)
    student = StudentsInfo.objects.filter(password_changed=False)
    for row_no, x in enumerate(student):
        credentials.write(row_no + 1, 0, x.user.username)
        credentials.write(row_no + 1, 1, x.first_password)
    wb.close()
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response[
        "Content-Disposition"
    ] = "attachment; filename=Student Login Credentials.xlsx"
    return response
