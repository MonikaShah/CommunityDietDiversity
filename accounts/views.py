import io
import ast
import openpyxl
import string
import random
import xlsxwriter
from datetime import datetime, date
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib import messages
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.models import Group, User
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.password_validation import validate_password
from .decorators import *
from .models import *
from .forms import *
from .bulkRegValidator import *
from shared.encryption import EncryptionHelper

encryptionHelper = EncryptionHelper()


def random_password_generator():
    return "".join(random.choices(string.ascii_lowercase + string.digits, k=8))


def valid_date(date):
    date_year = int(date[6:])
    date_month = int(date[:2])
    date_date = int(date[3:5])
    try:
        datetime(date_year, date_month, date_date)
        return True
    except:
        return False


def is_adult_func(dob):
    today = str(date.today())
    user_dob_year = int(dob[6:])
    user_dob_month = int(dob[:2])
    user_dob_date = int(dob[3:5])
    today_year = int(today[:4])
    today_month = int(today[5:7])
    today_date = int(today[8:])
    is_adult = False
    if (today_year - 18 > user_dob_year) or (
        (today_year - 18 == user_dob_year)
        and (
            (today_month > user_dob_month)
            or ((today_month == user_dob_month) and (today_date >= user_dob_date))
        )
    ):
        is_adult = True
    return str(is_adult)


def root(request):
    return redirect("accounts:loginlink")


@registration_data_cleanup
@redirect_to_dashboard
def registration(request):
    if request.method == "GET":
        form = RegistrationForm()
        return render(request, "registration/registration.html", {"form": form})
    else:
        form = RegistrationForm(request.POST)
        if form.is_valid():
            dob = request.POST["dob"]
            request.session["dob"] = dob
            request.session["registration_visited"] = True
            if is_adult_func(dob) == "True":
                return redirect("accounts:consent_adult")
            else:
                return redirect("accounts:consent")
        else:
            return render(request, "registration/registration.html", {"form": form})


@redirect_to_dashboard
def consent_adult(request):
    if "registration_visited" not in request.session:
        return redirect("accounts:registration")
    if request.method == "GET":
        if request.session.get("consent_data"):
            form = ConsentForm(request.session.get("consent_data"))
        else:
            form = ConsentForm()
        return render(
            request, "registration/consent.html", {"form": form, "is_adult": True}
        )
    else:
        form = ConsentForm(request.POST)
        if form.is_valid():
            request.session["consent_visited"] = True
            request.session["consent_data"] = request.POST
            return redirect("accounts:students_info_adult")
        else:
            return render(
                request, "registration/consent.html", {"form": form, "is_adult": True}
            )


@redirect_to_dashboard
def students_info_adult(request):
    return students_info(request, True)


@redirect_to_dashboard
def consent(request):
    if "registration_visited" not in request.session:
        return redirect("accounts:registration")
    if request.method == "GET":
        if request.session.get("consent_data"):
            form = ConsentForm(request.session.get("consent_data"))
        else:
            form = ConsentForm()
        return render(
            request, "registration/consent.html", {"form": form, "is_adult": False}
        )
    else:
        form = ConsentForm(request.POST)
        if form.is_valid():
            request.session["consent_visited"] = True
            request.session["consent_data"] = request.POST
            return redirect("accounts:students_info")
        else:
            return render(
                request, "registration/consent.html", {"form": form, "is_adult": False}
            )


@redirect_to_dashboard
def parents_info(request):
    if "registration_visited" not in request.session:
        return redirect("accounts:registration")
    if "consent_visited" not in request.session:
        return redirect("accounts:consent")
    if request.method == "GET":
        if request.session.get("data"):
            form = ParentsInfoForm(request.session.get("data"))
            user_creation_form = UserCreationForm(request.session.get("data"))
        else:
            form = ParentsInfoForm()
            user_creation_form = UserCreationForm()
        return render(
            request,
            "registration/parents_info.html",
            {
                "form": form,
                "user_creation_form": user_creation_form,
                "valid_state": True,
                "valid_city": True,
            },
        )
    else:
        form = ParentsInfoForm(request.POST)
        user_creation_form = UserCreationForm(request.POST)

        if form.is_valid() and user_creation_form.is_valid():
            temp = check_state_city(True, 0, str(request.POST["state"]))
            if temp[0]:
                if not check_state_city(False, temp[1], str(request.POST["city"])):
                    return render(
                        request,
                        "registration/parents_info.html",
                        {
                            "form": form,
                            "user_creation_form": user_creation_form,
                            "valid_state": True,
                            "valid_city": False,
                        },
                    )
            else:
                return render(
                    request,
                    "registration/parents_info.html",
                    {
                        "form": form,
                        "user_creation_form": user_creation_form,
                        "valid_state": False,
                        "valid_city": True,
                    },
                )
            request.session["data"] = request.POST
            request.session["parents_info_visited"] = True
            return redirect("accounts:students_info")
        else:
            return render(
                request,
                "registration/parents_info.html",
                {
                    "form": form,
                    "user_creation_form": user_creation_form,
                    "valid_state": True,
                    "valid_city": True,
                },
            )


@redirect_to_dashboard
def students_info(request, is_adult=False):
    if is_adult:
        if "registration_visited" not in request.session:
            return redirect("accounts:registration")
        if "consent_visited" not in request.session:
            return redirect("accounts:consent_adult")
    else:
        if "registration_visited" not in request.session:
            return redirect("accounts:registration")
        if "consent_visited" not in request.session:
            return redirect("accounts:consent")
        if "parents_info_visited" not in request.session:
            return redirect("accounts:parents_info")
    student_dob = request.session["dob"]
    if request.method == "GET":
        form = StudentsInfoForm()
        user_creation_form = UserCreationForm()
        return render(
            request,
            "registration/students_info.html",
            {
                "form": form,
                "user_creation_form": user_creation_form,
                "is_adult": is_adult,
                "student_dob": student_dob,
                "valid_state": True,
                "valid_city": True,
            },
        )
    else:
        if not is_adult:
            previousPOST = request.session["data"]
            form = StudentsInfoForm(request.POST)
            studentuserform = UserCreationForm(request.POST)
            parentform = ParentsInfoForm(previousPOST)
            parentuserform = UserCreationForm(previousPOST)
            if form.is_valid() and studentuserform.is_valid():
                temp = check_state_city(True, 0, str(request.POST["state"]))
                if temp[0]:
                    if not check_state_city(False, temp[1], str(request.POST["city"])):
                        return render(
                            request,
                            "registration/students_info.html",
                            {
                                "form": form,
                                "user_creation_form": studentuserform,
                                "is_adult": is_adult,
                                "student_dob": student_dob,
                                "valid_state": True,
                                "valid_city": False,
                            },
                        )
                else:
                    return render(
                        request,
                        "registration/students_info.html",
                        {
                            "form": form,
                            "user_creation_form": studentuserform,
                            "is_adult": is_adult,
                            "student_dob": student_dob,
                            "valid_state": False,
                            "valid_city": True,
                        },
                    )
                parentUser = parentuserform.save(commit=False)
                parentUser.save()
                parent_group = Group.objects.get(name="Parents")
                parentUser.groups.add(parent_group)
                parentUser.save()

                parent = parentform.save(commit=False)
                parent.user = parentUser
                parent.email = encryptionHelper.encrypt(previousPOST["email"])
                parent.name = encryptionHelper.encrypt(previousPOST["name"])
                parent.dob = encryptionHelper.encrypt(previousPOST["dob"])
                parent.mobile_no = encryptionHelper.encrypt(previousPOST["mobile_no"])
                parent.gender = encryptionHelper.encrypt(previousPOST["gender"])
                parent.state = State.objects.get(
                    state__icontains=previousPOST["state"].strip()
                )
                parent.city = City.objects.get(
                    city__icontains=previousPOST["city"].strip()
                )
                parent.address = encryptionHelper.encrypt(previousPOST["address"])
                parent.pincode = encryptionHelper.encrypt(previousPOST["pincode"])
                parent.no_of_family_members = encryptionHelper.encrypt(
                    previousPOST["no_of_family_members"]
                )
                parent.children_count = encryptionHelper.encrypt(
                    previousPOST["children_count"]
                )
                parent.save()

                studentuser = studentuserform.save(commit=False)
                studentuser.save()
                student_group = Group.objects.get(name="Students")
                studentuser.groups.add(student_group)
                studentuser.save()

                student = form.save(commit=False)
                student.user = studentuser
                student.rollno = encryptionHelper.encrypt(request.POST["rollno"])
                student.name = encryptionHelper.encrypt(request.POST["name"])
                student.email = encryptionHelper.encrypt(request.POST["email"])
                student.dob = encryptionHelper.encrypt(student_dob)
                student.mobile_no = encryptionHelper.encrypt(request.POST["mobile_no"])
                student.gender = encryptionHelper.encrypt(request.POST["gender"])
                student.adult = encryptionHelper.encrypt(is_adult_func(student_dob))
                student.state = State.objects.get(
                    state__icontains=request.POST["state"].strip()
                )
                student.city = City.objects.get(
                    city__icontains=request.POST["city"].strip()
                )
                student.pincode = encryptionHelper.encrypt(request.POST["pincode"])
                student.address = encryptionHelper.encrypt(request.POST["address"])
                student.parent = parent
                student.save()

                user = authenticate(
                    request,
                    username=previousPOST["username"],
                    password=previousPOST["password1"],
                )
                request.session.set_expiry(86400)
                if user is not None:
                    login(request, user)

                del request.session["consent_data"]
                del request.session["data"]
                del request.session["dob"]
                del request.session["registration_visited"]
                del request.session["consent_visited"]
                del request.session["parents_info_visited"]
                return redirect("accounts:parent_dashboard")
            else:
                return render(
                    request,
                    "registration/students_info.html",
                    {
                        "form": form,
                        "user_creation_form": studentuserform,
                        "is_adult": is_adult,
                        "student_dob": student_dob,
                        "valid_state": True,
                        "valid_city": True,
                    },
                )
        else:
            form = StudentsInfoForm(request.POST)
            studentuserform = UserCreationForm(request.POST)
            if form.is_valid() and studentuserform.is_valid():
                temp = check_state_city(True, 0, str(request.POST["state"]))
                if temp[0]:
                    if not check_state_city(False, temp[1], str(request.POST["city"])):
                        return render(
                            request,
                            "registration/students_info.html",
                            {
                                "form": form,
                                "user_creation_form": studentuserform,
                                "is_adult": is_adult,
                                "student_dob": student_dob,
                                "valid_state": True,
                                "valid_city": False,
                            },
                        )
                else:
                    return render(
                        request,
                        "registration/students_info.html",
                        {
                            "form": form,
                            "user_creation_form": studentuserform,
                            "is_adult": is_adult,
                            "student_dob": student_dob,
                            "valid_state": False,
                            "valid_city": True,
                        },
                    )
                studentuser = studentuserform.save(commit=False)
                studentuser.save()
                student_group = Group.objects.get(name="Students")
                studentuser.groups.add(student_group)
                studentuser.save()

                student = form.save(commit=False)
                student.user = studentuser
                student.rollno = encryptionHelper.encrypt(request.POST["rollno"])
                student.name = encryptionHelper.encrypt(request.POST["name"])
                student.email = encryptionHelper.encrypt(request.POST["email"])
                student.dob = encryptionHelper.encrypt(student_dob)
                student.mobile_no = encryptionHelper.encrypt(request.POST["mobile_no"])
                student.gender = encryptionHelper.encrypt(request.POST["gender"])
                student.adult = encryptionHelper.encrypt(is_adult_func(student_dob))
                student.state = State.objects.get(
                    state__icontains=request.POST["state"].strip()
                )
                student.city = City.objects.get(
                    city__icontains=request.POST["city"].strip()
                )
                student.pincode = encryptionHelper.encrypt(request.POST["pincode"])
                student.address = encryptionHelper.encrypt(request.POST["address"])
                student.save()

                user = authenticate(
                    request,
                    username=request.POST["username"],
                    password=request.POST["password1"],
                )
                request.session.set_expiry(86400)
                if user is not None:
                    login(request, user)

                del request.session["consent_data"]
                del request.session["dob"]
                del request.session["registration_visited"]
                del request.session["consent_visited"]
                return redirect("accounts:student_dashboard")
            else:
                return render(
                    request,
                    "registration/students_info.html",
                    {
                        "form": form,
                        "user_creation_form": studentuserform,
                        "is_adult": is_adult,
                        "student_dob": student_dob,
                        "valid_state": True,
                        "valid_city": True,
                    },
                )


@registration_data_cleanup
@redirect_to_dashboard
def loginU(request):
    if request.method == "GET":
        form = CustomAuthenticationForm()
        return render(request, "registration/login.html", {"form": form})
    else:
        username = request.POST["username"]
        password = request.POST["password"]
        grp = request.POST["groups"]
        user = authenticate(request, username=username, password=password)
        request.session.set_expiry(86400)
        grp_name = Group.objects.get(pk=grp).name
        form = CustomAuthenticationForm(request.POST)
        if user is not None:
            if is_member(user, grp):
                login(request, user)
                if grp_name == "Parents":
                    return redirect("accounts:parent_dashboard")
                elif grp_name == "Students":
                    return redirect("accounts:student_dashboard")
                elif grp_name == "Teachers":
                    return redirect("accounts:teacher_all_sessions")
                elif grp_name == "Coordinators":
                    return redirect("accounts:coordinator_dashboard")
                elif grp_name == "Super Coordinators":
                    return redirect("accounts:supercoordinator_dashboard")
            else:
                messages.error(request, "User does not belong to selected group")
                return render(request, "registration/login.html", {"form": form})
        else:
            messages.error(request, "Invalid credentials")
            return render(request, "registration/login.html", {"form": form})


@login_required(login_url="accounts:loginlink")
def logoutU(request):
    logout(request)
    return redirect("accounts:loginlink")


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_parent, login_url="accounts:forbidden")
@password_change_required
def parent_dashboard(request):
    students = (
        ParentsInfo.objects.filter(user=request.user).first().studentsinfo_set.all()
    )
    for student in students:
        student.name = encryptionHelper.decrypt(student.name)
    return render(
        request,
        "parent/parent_dashboard.html",
        {"students": students, "page_type": "parent_dashboard"},
    )


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_admin, login_url="accounts:forbidden")
def addSuperCoordinatorForm(request):
    if request.method == "GET":
        form = SuperCoordinatorsInfoForm()
        user_creation_form = UserCreationForm()
        return render(
            request,
            "admin/add_supercoordinator.html",
            {"form": form, "user_creation_form": user_creation_form},
        )
    else:
        form = SuperCoordinatorsInfoForm(request.POST)
        supercoordinatoruserform = UserCreationForm(request.POST)
        if form.is_valid() and supercoordinatoruserform.is_valid():
            supercoordinatoruser = supercoordinatoruserform.save(commit=False)
            supercoordinatoruser.save()
            supercoordinator_group = Group.objects.get(name="Super Coordinators")
            supercoordinatoruser.groups.add(supercoordinator_group)
            supercoordinatoruser.save()
            supercoordinator = form.save(commit=False)
            supercoordinator.user = supercoordinatoruser
            supercoordinator.email = encryptionHelper.encrypt(request.POST["email"])
            supercoordinator.name = encryptionHelper.encrypt(request.POST["name"])
            supercoordinator.dob = encryptionHelper.encrypt(request.POST["dob"])
            supercoordinator.gender = encryptionHelper.encrypt(request.POST["gender"])
            supercoordinator.mobile_no = encryptionHelper.encrypt(
                request.POST["mobile_no"]
            )
            supercoordinator.save()
            return redirect("accounts:add_supercoordinator_form")
        else:
            return render(
                request,
                "admin/add_supercoordinator.html",
                {"form": form, "user_creation_form": supercoordinatoruserform},
            )


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_parent, login_url="accounts:forbidden")
@password_change_required
def addStudentForm(request):
    if request.method == "GET":
        form = StudentsInfoForm()
        user_creation_form = UserCreationForm()
        return render(
            request,
            "parent/add_student.html",
            {
                "form": form,
                "user_creation_form": user_creation_form,
                "valid_state": True,
                "valid_city": True,
            },
        )
    else:
        form = StudentsInfoForm(request.POST)
        studentuserform = UserCreationForm(request.POST)
        if form.is_valid() and studentuserform.is_valid():
            temp = check_state_city(True, 0, str(request.POST["state"]))
            if temp[0]:
                if not check_state_city(False, temp[1], str(request.POST["city"])):
                    return render(
                        request,
                        "parent/add_student.html",
                        {
                            "form": form,
                            "user_creation_form": studentuserform,
                            "valid_state": True,
                            "valid_city": False,
                        },
                    )
            else:
                return render(
                    request,
                    "parent/add_student.html",
                    {
                        "form": form,
                        "user_creation_form": studentuserform,
                        "valid_state": False,
                        "valid_city": True,
                    },
                )
            studentuser = studentuserform.save(commit=False)
            studentuser.save()
            student_group = Group.objects.get(name="Students")
            studentuser.groups.add(student_group)
            studentuser.save()
            student = form.save(commit=False)
            student.user = studentuser
            student.rollno = encryptionHelper.encrypt(request.POST["rollno"])
            student.name = encryptionHelper.encrypt(request.POST["name"])
            student.email = encryptionHelper.encrypt(request.POST["email"])
            student.dob = encryptionHelper.encrypt(request.POST["dob"])
            student.mobile_no = encryptionHelper.encrypt(request.POST["mobile_no"])
            student.gender = encryptionHelper.encrypt(request.POST["gender"])
            student.adult = encryptionHelper.encrypt(is_adult_func(request.POST["dob"]))
            student.state = State.objects.get(
                state__icontains=request.POST["state"].strip()
            )
            student.city = City.objects.get(
                city__icontains=request.POST["city"].strip()
            )
            student.pincode = encryptionHelper.encrypt(request.POST["pincode"])
            student.address = encryptionHelper.encrypt(request.POST["address"])
            student.parent = ParentsInfo.objects.filter(user=request.user).first()
            student.first_password = ""
            student.password_changed = True
            student.save()
            return redirect("accounts:parent_dashboard")
        else:
            return render(
                request,
                "parent/add_student.html",
                {
                    "form": form,
                    "user_creation_form": studentuserform,
                    "valid_state": True,
                    "valid_city": True,
                },
            )


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_coordinator, login_url="accounts:forbidden")
@password_change_required
def addTeacherForm(request):
    if request.method == "GET":
        form = TeachersInfoForm()
        user_creation_form = UserCreationForm()
        return render(
            request,
            "coordinator/add_teacher.html",
            {"form": form, "user_creation_form": user_creation_form},
        )
    else:
        form = TeachersInfoForm(request.POST)
        teacheruserform = UserCreationForm(request.POST)
        if form.is_valid() and teacheruserform.is_valid():
            teacheruser = teacheruserform.save(commit=False)
            teacheruser.save()
            teacher_group = Group.objects.get(name="Teachers")
            teacheruser.groups.add(teacher_group)
            teacheruser.save()
            teacher = form.save(commit=False)
            teacher.user = teacheruser
            teacher.email = encryptionHelper.encrypt(request.POST["email"])
            teacher.name = encryptionHelper.encrypt(request.POST["name"])
            teacher.dob = encryptionHelper.encrypt(request.POST["dob"])
            teacher.gender = encryptionHelper.encrypt(request.POST["gender"])
            teacher.mobile_no = encryptionHelper.encrypt(request.POST["mobile_no"])
            teacher.organization = Organization.objects.filter(
                name=CoordinatorInCharge.objects.filter(user=request.user)
                .first()
                .organization
            ).first()
            teacher.coordinator = CoordinatorInCharge.objects.filter(
                user=request.user
            ).first()
            teacher.save()
            return redirect("accounts:coordinator_dashboard")
        else:
            return render(
                request,
                "coordinator/add_teacher.html",
                {"form": form, "user_creation_form": teacheruserform},
            )


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_supercoordinator, login_url="accounts:forbidden")
def addCoordinatorForm(request, id):
    if request.method == "GET":
        form = CoordinatorsInfoForm()
        user_creation_form = UserCreationForm()
        return render(
            request,
            "supercoordinator/add_coordinator.html",
            {"form": form, "user_creation_form": user_creation_form},
        )
    else:
        form = CoordinatorsInfoForm(request.POST)
        coordinatoruserform = UserCreationForm(request.POST)
        if form.is_valid() and coordinatoruserform.is_valid():
            coordinatoruser = coordinatoruserform.save(commit=False)
            coordinatoruser.save()
            coordinator_group = Group.objects.get(name="Coordinators")
            coordinatoruser.groups.add(coordinator_group)
            coordinatoruser.save()
            coordinator = form.save(commit=False)
            coordinator.user = coordinatoruser
            coordinator.email = encryptionHelper.encrypt(request.POST["email"])
            coordinator.name = encryptionHelper.encrypt(request.POST["name"])
            coordinator.dob = encryptionHelper.encrypt(request.POST["dob"])
            coordinator.gender = encryptionHelper.encrypt(request.POST["gender"])
            coordinator.mobile_no = encryptionHelper.encrypt(request.POST["mobile_no"])
            coordinator.super_coordinator = SuperCoordinator.objects.filter(
                user=request.user
            ).first()
            coordinator.organization = Organization.objects.filter(id=id).first()
            coordinator.save()
            return redirect("accounts:view_coordinators", id)
        else:
            return render(
                request,
                "supercoordinator/add_coordinator.html",
                {"form": form, "user_creation_form": coordinatoruserform},
            )


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_supercoordinator, login_url="accounts:forbidden")
def addOrganizationForm(request):
    if request.method == "GET":
        form = OrganizationsInfoForm()
        return render(
            request,
            "supercoordinator/add_organization.html",
            {
                "form": form,
                "valid_state": True,
                "valid_city": True,
            },
        )
    else:
        form = OrganizationsInfoForm(request.POST)
        if form.is_valid():
            temp = check_state_city(True, 0, str(request.POST["state"]))
            if temp[0]:
                if not check_state_city(False, temp[1], str(request.POST["city"])):
                    return render(
                        request,
                        "supercoordinator/add_organization.html",
                        {
                            "form": form,
                            "valid_state": True,
                            "valid_city": False,
                        },
                    )
            else:
                return render(
                    request,
                    "supercoordinator/add_organization.html",
                    {
                        "form": form,
                        "valid_state": False,
                        "valid_city": True,
                    },
                )
            organization = form.save(commit=False)
            organization.state = State.objects.get(
                state__icontains=request.POST["state"].strip()
            )
            organization.city = City.objects.get(
                city__icontains=request.POST["city"].strip()
            )
            organization.save()
            return redirect("accounts:supercoordinator_dashboard")
        else:
            return render(
                request,
                "supercoordinator/add_organization.html",
                {"form": form},
            )


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_coordinator, login_url="accounts:forbidden")
@password_change_required
def addSessionForm(request):
    if request.method == "GET":
        form = SessionsInfoForm()
        return render(
            request,
            "coordinator/add_session.html",
            {"form": form},
        )
    else:
        form = SessionsInfoForm(request.POST)
        if form.is_valid():
            session = form.save(commit=False)
            session.coordinator = CoordinatorInCharge.objects.filter(
                user=request.user
            ).first()
            session.save()
            return redirect("accounts:all_sessions")
        else:
            return render(
                request,
                "coordinator/add_session.html",
                {"form": form},
            )


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_coordinator, login_url="accounts:forbidden")
@password_change_required
def viewSessionTeachers(request, id, my_messages=None):
    session = Session.objects.filter(id=id).first()
    objects = Teacher_Session.objects.filter(session=session)
    teachers = []
    for object in objects:
        object.teacher.name = encryptionHelper.decrypt(object.teacher.name)
        teachers.append(object.teacher)
    if my_messages != None:
        return render(
            request,
            "coordinator/view_session_teachers.html",
            {
                "teachers": teachers,
                "session": session,
                "my_messages": my_messages,
                "page_type": "view_session_teachers",
            },
        )
    else:
        return render(
            request,
            "coordinator/view_session_teachers.html",
            {
                "teachers": teachers,
                "session": session,
                "page_type": "view_session_teachers",
            },
        )


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_coordinator, login_url="accounts:forbidden")
@password_change_required
def getSessionTeachersTemplate(request):
    output = io.BytesIO()
    wb = xlsxwriter.Workbook(output)
    ws = wb.add_worksheet("Session Teachers Data")
    columns = [
        "Teacher Username",
    ]
    for col_num in range(len(columns)):
        ws.write(0, col_num, columns[col_num])
    wb.close()
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = "attachment; filename=sessionTeacherTemplate.xlsx"
    return response


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_coordinator, login_url="accounts:forbidden")
@password_change_required
def addSessionTeachers(request, id):
    if request.method == "GET":
        return render(
            request,
            "coordinator/add_session_teachers.html",
            {"page_type": "add_session_teachers"},
        )
    else:
        try:
            excel_file = request.FILES["excel_file"]
            if excel_file.name[-4:] == "xlsx":
                try:
                    wb = openpyxl.load_workbook(excel_file)
                    teacherSheet = wb["Session Teachers Data"]
                except:
                    return render(
                        request,
                        "coordinator/add_session_teachers.html",
                        {
                            "page_type": "add_session_teachers",
                            "my_messages": {
                                "error": "Incorrect file uploaded, please use the template provided above."
                            },
                        },
                    )
            else:
                return render(
                    request,
                    "coordinator/add_session_teachers.html",
                    {
                        "page_type": "add_session_teachers",
                        "my_messages": {
                            "error": "Incorrect file type, only .xlsx files are allowed!"
                        },
                    },
                )
        except:
            return render(
                request,
                "coordinator/add_session_teachers.html",
                {
                    "page_type": "add_session_teachers",
                    "my_messages": {"error": "Sorry something went wrong!"},
                },
            )

        organization = (
            CoordinatorInCharge.objects.filter(user=request.user).first().organization
        )
        breaking = error = False
        error_message = ""
        teacher_data = []
        for row_no, row in enumerate(teacherSheet.iter_rows()):
            if breaking == True:
                break
            if row_no == 0:
                continue
            for cell in row:
                if cell.column_letter == "A":
                    if cell.value == None:
                        breaking = True
                        break
                    else:
                        if User.objects.filter(username=cell.value).exists():
                            user = User.objects.filter(username=cell.value)
                            if is_teacher(user):
                                teacher_user = TeacherInCharge.objects.filter(
                                    user=user
                                ).first()
                                if teacher_user.organization == organization:
                                    teacher_data.append(teacher_user)
                                else:
                                    breaking = error = True
                                    error_message = (
                                        "Teacher at row number "
                                        + str(row_no + 1)
                                        + " does not belong to your organization"
                                    )
                                    break
                            else:
                                breaking = error = True
                                error_message = (
                                    "User at row number "
                                    + str(row_no + 1)
                                    + " is not a teacher"
                                )
                                break
                        else:
                            breaking = error = True
                            error_message = "Invalid username at row number " + str(
                                row_no + 1
                            )
                            break

        if error == True:
            return render(
                request,
                "coordinator/add_session_teachers.html",
                {
                    "page_type": "add_session_teachers",
                    "my_messages": {"error": error_message},
                },
            )

        session = Session.objects.filter(id=id).first()
        teacher_session = Teacher_Session.objects.filter(session=session)
        teacher_session_teacher_objects = []
        for i in teacher_session:
            teacher_session_teacher_objects.append(i.teacher)
        teacher_added = 0
        teacher_exists = 0
        for teacher_user in teacher_data:
            if teacher_user in teacher_session_teacher_objects:
                teacher_exists += 1
            else:
                new_teacher_session = Teacher_Session()
                new_teacher_session.session = session
                new_teacher_session.teacher = teacher_user
                new_teacher_session.save()
                teacher_added += 1

        if teacher_added == 0:
            my_messages = {
                "success": "Registration successful. Already registered: "
                + str(teacher_exists)
            }
        elif teacher_exists == 0:
            my_messages = {
                "success": "Registration successful. Newly registered: "
                + str(teacher_added)
            }
        else:
            my_messages = {
                "success": "Registration successful. Already registered: "
                + str(teacher_exists)
                + ", Newly registered: "
                + str(teacher_added)
            }

        return viewSessionTeachers(request, id, my_messages)


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_teacher, login_url="accounts:forbidden")
@password_change_required
def viewSessionStudents(request, id, my_messages=None):
    session = Session.objects.filter(id=id).first()
    teacher = TeacherInCharge.objects.filter(user=request.user).first()
    students = StudentsInfo.objects.filter(session=session, teacher=teacher)
    for student in students:
        student.name = encryptionHelper.decrypt(student.name)
        student.rollno = encryptionHelper.decrypt(student.rollno)
    if my_messages != None:
        return render(
            request,
            "teacher/view_session_students.html",
            {
                "students": students,
                "session": session,
                "my_messages": my_messages,
                "page_type": "view_session_students",
            },
        )
    else:
        return render(
            request,
            "teacher/view_session_students.html",
            {
                "students": students,
                "session": session,
                "page_type": "view_session_students",
            },
        )


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_teacher, login_url="accounts:forbidden")
@password_change_required
def getSessionStudentsTemplate(request):
    output = io.BytesIO()
    wb = xlsxwriter.Workbook(output)
    ws = wb.add_worksheet("Session Students Data")
    columns = [
        "Student Username",
    ]
    for col_num in range(len(columns)):
        ws.write(0, col_num, columns[col_num])
    wb.close()
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = "attachment; filename=sessionStudentTemplate.xlsx"
    return response


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_teacher, login_url="accounts:forbidden")
@password_change_required
def addSessionStudents(request, id):
    if request.method == "GET":
        return render(
            request,
            "teacher/add_session_students.html",
            {"page_type": "add_session_students"},
        )
    else:
        try:
            excel_file = request.FILES["excel_file"]
            if excel_file.name[-4:] == "xlsx":
                try:
                    wb = openpyxl.load_workbook(excel_file)
                    studentSheet = wb["Session Students Data"]
                except:
                    return render(
                        request,
                        "teacher/add_session_students.html",
                        {
                            "page_type": "add_session_students",
                            "my_messages": {
                                "error": "Incorrect file uploaded, please use the template provided above."
                            },
                        },
                    )
            else:
                return render(
                    request,
                    "teacher/add_session_students.html",
                    {
                        "page_type": "add_session_students",
                        "my_messages": {
                            "error": "Incorrect file type, only .xlsx files are allowed!"
                        },
                    },
                )
        except:
            return render(
                request,
                "teacher/add_session_students.html",
                {
                    "page_type": "add_session_students",
                    "my_messages": {"error": "Sorry something went wrong!"},
                },
            )

        organization = (
            TeacherInCharge.objects.filter(user=request.user).first().organization
        )
        breaking = error = False
        error_message = ""
        student_data = []
        for row_no, row in enumerate(studentSheet.iter_rows()):
            if breaking == True:
                break
            if row_no == 0:
                continue
            for cell in row:
                if cell.column_letter == "A":
                    if cell.value == None:
                        breaking = True
                        break
                    else:
                        if User.objects.filter(username=cell.value).exists():
                            user = User.objects.filter(username=cell.value)
                            if is_student(user):
                                student_user = StudentsInfo.objects.filter(
                                    user=user
                                ).first()
                                if student_user.organization == organization:
                                    student_data.append(student_user)
                                else:
                                    breaking = error = True
                                    error_message = (
                                        "Student at row number "
                                        + str(row_no + 1)
                                        + " does not belong to your organization"
                                    )
                                    break
                            else:
                                breaking = error = True
                                error_message = (
                                    "User at row number "
                                    + str(row_no + 1)
                                    + " is not a student"
                                )
                                break
                        else:
                            breaking = error = True
                            error_message = "Invalid username at row number " + str(
                                row_no + 1
                            )
                            break

        if error == True:
            return render(
                request,
                "teacher/add_session_students.html",
                {
                    "page_type": "add_session_students",
                    "my_messages": {"error": error_message},
                },
            )

        session = Session.objects.filter(id=id).first()
        student_added = 0
        student_exists = 0
        for student_user in student_data:
            if student_user.session == session:
                student_exists += 1
            else:
                student_user.session = session
                student_user.save()
                student_added += 1

        if student_added == 0:
            my_messages = {
                "success": "Registration successful. Already registered: "
                + str(student_exists)
            }
        elif student_exists == 0:
            my_messages = {
                "success": "Registration successful. Newly registered: "
                + str(student_added)
            }
        else:
            my_messages = {
                "success": "Registration successful. Already registered: "
                + str(student_exists)
                + ", Newly registered: "
                + str(student_added)
            }

        return viewSessionStudents(request, id, my_messages)


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_teacher, login_url="accounts:forbidden")
@password_change_required
def getTemplate(request):
    output = io.BytesIO()
    wb = xlsxwriter.Workbook(output)
    align = wb.add_format()
    align.set_align("center")
    bold = wb.add_format({"bold": True})
    ws = wb.add_worksheet("Parents Data")
    ws2 = wb.add_worksheet("Students Data")

    columns = [
        "parentId",
        "Parent Username",
        "Parent Name",
        "Email id",
        "Phone Number",
        "Gender",
        "Date of Birth",
        "Pincode",
        "State",
        "City",
        "Address",
        "Religion",
        "Type of family",
        "No of family members",
        "Children Count",
        "Education",
        "Occupation",
    ]
    columns2 = [
        "Student Name",
        "Email id",
        "Phone Number",
        "Gender",
        "Date of Birth",
        "Pincode",
        "State",
        "City",
        "Address",
        "Roll Number",
        "parentId",
    ]
    for col_num in range(len(columns)):
        ws.write(0, col_num, columns[col_num], bold)
    for col_num in range(len(columns2)):
        ws2.write(0, col_num, columns2[col_num], bold)
    for row_num in range(1, 1000):
        ws.write(row_num, 0, row_num)
        ws.write(row_num, 1, "-", align)

    wb.close()

    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response[
        "Content-Disposition"
    ] = "attachment; filename=bulkRegistrationTemplate.xlsx"
    return response


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_teacher, login_url="accounts:forbidden")
@password_change_required
def bulkRegister(request):
    if request.method == "GET":
        return render(
            request,
            "teacher/bulkregistration.html",
            {"page_type": "bulk_register"},
        )
    else:
        try:
            excel_file = request.FILES["excel_file"]
            if excel_file.name[-4:] == "xlsx":
                try:
                    wb = openpyxl.load_workbook(excel_file)
                    parentSheet = wb["Parents Data"]
                    studentSheet = wb["Students Data"]
                except:
                    return render(
                        request,
                        "teacher/bulkregistration.html",
                        {
                            "page_type": "bulk_register",
                            "my_messages": {
                                "error": "Incorrect file uploaded, please use the template provided above."
                            },
                        },
                    )
            else:
                return render(
                    request,
                    "teacher/bulkregistration.html",
                    {
                        "page_type": "bulk_register",
                        "my_messages": {
                            "error": "Incorrect file type, only .xlsx files are allowed!"
                        },
                    },
                )
        except:
            return render(
                request,
                "teacher/bulkregistration.html",
                {
                    "page_type": "bulk_register",
                    "my_messages": {"error": "Sorry something went wrong!"},
                },
            )

        breaking = error = False
        error_message = ""
        parent_data = []
        for row_no, row in enumerate(parentSheet.iter_rows()):
            if breaking == True:
                break
            if row_no == 0:
                continue
            row_data = []
            for cell in row:
                if cell.column_letter == "A":
                    if cell.value == None:
                        breaking = True
                        break
                    elif cell.value != row_no:
                        breaking = error = True
                        error_message = (
                            "There something wrong in the parentId column in parent data sheet at row number "
                            + str(row_no + 1)
                        )
                        break
                elif cell.column_letter == "B":
                    if cell.value == "-" or cell.value == None:
                        row_data.append(None)
                    else:
                        if User.objects.filter(username=str(cell.value)).exists():
                            row_data.append(cell.value)
                            break
                        else:
                            breaking = error = True
                            error_message = (
                                "Parent user with the following username at row number "
                                + str(row_no + 1)
                                + " does not exist"
                            )
                            break
                elif cell.column_letter == "C":
                    if cell.value == None:
                        breaking = True
                        break
                    if not valid_name(cell.value):
                        breaking = error = True
                        error_message = "Invalid parent name at row number " + str(
                            row_no + 1
                        )
                        break
                    else:
                        row_data.append(cell.value)
                elif cell.column_letter == "D":
                    if cell.value != None and not valid_email(cell.value):
                        breaking = error = True
                        error_message = "Invalid parent email at row number " + str(
                            row_no + 1
                        )
                        break
                    else:
                        row_data.append(cell.value)
                elif cell.column_letter == "E":
                    if cell.value != None and not valid_mobile_no(cell.value):
                        breaking = error = True
                        error_message = (
                            "Invalid parent mobile number at row number "
                            + str(row_no + 1)
                        )
                        break
                    else:
                        row_data.append(cell.value)
                elif cell.column_letter == "F":
                    if cell.value == None:
                        breaking = error = True
                        error_message = "Parent's gender missing at row number " + str(
                            row_no + 1
                        )
                        break
                    else:
                        if check_gender(str(cell.value)):
                            row_data.append(cell.value)
                        else:
                            breaking = error = True
                            error_message = (
                                "Invalid gender input for parents data at row number  "
                                + str(row_no + 1)
                                + ", please read the instructions"
                            )
                            break
                elif cell.column_letter == "G":
                    if cell.value == None:
                        breaking = error = True
                        error_message = (
                            "Parent's date of birth missing at row number "
                            + str(row_no + 1)
                        )
                        break
                    else:
                        if re.match(
                            "^[0-9]{2}/[0-9]{2}/[0-9]{4}$", str(cell.value)
                        ) and valid_date(str(cell.value)):
                            if is_adult_func(str(cell.value)) == "True":
                                row_data.append(cell.value)
                            else:
                                breaking = error = True
                                error_message = (
                                    "Parent at row number "
                                    + str(row_no + 1)
                                    + " is not an adult"
                                )
                            break
                        else:
                            breaking = error = True
                            error_message = (
                                "Invalid parent's date of birth at row number "
                                + str(row_no + 1)
                            )
                            break
                elif cell.column_letter == "H":
                    if cell.value == None:
                        breaking = error = True
                        error_message = "Parent's pincode missing at row number " + str(
                            row_no + 1
                        )
                        break
                    else:
                        if not valid_pincode(cell.value):
                            breaking = error = True
                            error_message = (
                                "Invalid parent's pincode at row number "
                                + str(row_no + 1)
                            )
                            break
                        else:
                            row_data.append(cell.value)
                elif cell.column_letter == "I":
                    if cell.value == None:
                        breaking = error = True
                        error_message = "Parent's state missing at row number " + str(
                            row_no + 1
                        )
                        break
                    else:
                        temp = check_state_city(True, 0, str(cell.value))
                        if temp[0]:
                            row_data.append(cell.value)
                            row_data.append(temp[1])
                        else:
                            breaking = error = True
                            error_message = (
                                "Invalid parent's state at row number "
                                + str(row_no + 1)
                            )
                            break
                elif cell.column_letter == "J":
                    if cell.value == None:
                        breaking = error = True
                        error_message = "Parent's city missing at row number " + str(
                            row_no + 1
                        )
                        break
                    else:
                        if check_state_city(False, row_data[-1], str(cell.value)):
                            row_data.pop()
                            row_data.append(cell.value)
                        else:
                            breaking = error = True
                            error_message = (
                                "Invalid parent's city at row number " + str(row_no + 1)
                            )
                            break
                elif cell.column_letter == "K":
                    if cell.value == None:
                        breaking = error = True
                        error_message = "Parent's address missing at row number " + str(
                            row_no + 1
                        )
                        break
                    else:
                        row_data.append(cell.value)
                elif cell.column_letter == "L":
                    if cell.value == None:
                        breaking = error = True
                        error_message = (
                            "Parent's religion missing at row number " + str(row_no + 1)
                        )
                        break
                    else:
                        if check_religion(str(cell.value)):
                            row_data.append(cell.value)
                        else:
                            breaking = error = True
                            error_message = (
                                "Invalid parent's religion at row number "
                                + str(row_no + 1)
                            )
                            break
                elif cell.column_letter == "M":
                    if cell.value == None:
                        breaking = error = True
                        error_message = (
                            "Parent's type of family missing at row number "
                            + str(row_no + 1)
                        )
                        break
                    else:
                        if check_type_of_fam(str(cell.value)):
                            row_data.append(cell.value)
                        else:
                            breaking = error = True
                            error_message = (
                                "Invalid parent's type of family at row number "
                                + str(row_no + 1)
                            )
                            break
                elif cell.column_letter == "N":
                    if cell.value == None:
                        breaking = error = True
                        error_message = (
                            "Parent's no of family members missing at row number "
                            + str(row_no + 1)
                        )
                        break
                    else:
                        try:
                            int(cell.value)
                        except:
                            breaking = error = True
                            error_message = (
                                "Invalid parent's no of family members at row number "
                                + str(row_no + 1)
                            )
                            break
                        row_data.append(cell.value)
                elif cell.column_letter == "O":
                    if cell.value == None:
                        breaking = error = True
                        error_message = (
                            "Parent's children count missing at row number "
                            + str(row_no + 1)
                        )
                        break
                    else:
                        try:
                            int(cell.value)
                        except:
                            breaking = error = True
                            error_message = (
                                "Invalid parent's children count at row number "
                                + str(row_no + 1)
                            )
                            break
                        row_data.append(cell.value)
                elif cell.column_letter == "P":
                    if cell.value == None:
                        breaking = error = True
                        error_message = (
                            "Parent's education missing at row number "
                            + str(row_no + 1)
                        )
                        break
                    else:
                        if check_education(str(cell.value)):
                            row_data.append(cell.value)
                        else:
                            breaking = error = True
                            error_message = (
                                "Invalid parent's education at row number "
                                + str(row_no + 1)
                            )
                            break
                elif cell.column_letter == "Q":
                    if cell.value == None:
                        breaking = error = True
                        error_message = (
                            "Parent's occupation missing at row number "
                            + str(row_no + 1)
                        )
                        break
                    else:
                        if check_occupation(str(cell.value)):
                            row_data.append(cell.value)
                        else:
                            breaking = error = True
                            error_message = (
                                "Invalid parent's occupation at row number "
                                + str(row_no + 1)
                            )
                            break
            parent_data.append(row_data)

        if error == True:
            return render(
                request,
                "teacher/bulkregistration.html",
                {
                    "page_type": "bulk_register",
                    "my_messages": {"error": error_message},
                },
            )

        breaking = error = False
        error_message = ""
        student_data = []
        for row_no, row in enumerate(studentSheet.iter_rows()):
            if breaking == True:
                break
            if row_no == 0:
                continue
            row_data = []
            for cell in row:
                if cell.column_letter == "A":
                    if cell.value == None:
                        breaking = True
                        break
                    if not valid_name(cell.value):
                        breaking = error = True
                        error_message = "Invalid student name at row number " + str(
                            row_no + 1
                        )
                        break
                    else:
                        row_data.append(cell.value)
                elif cell.column_letter == "B":
                    if cell.value != None and not valid_email(cell.value):
                        breaking = error = True
                        error_message = "Invalid student email at row number " + str(
                            row_no + 1
                        )
                        break
                    row_data.append(cell.value)
                elif cell.column_letter == "C":
                    if cell.value != None and not valid_mobile_no(cell.value):
                        breaking = error = True
                        error_message = (
                            "Invalid student mobile number at row number "
                            + str(row_no + 1)
                        )
                        break
                    row_data.append(cell.value)
                elif cell.column_letter == "D":
                    if cell.value == None:
                        breaking = error = True
                        error_message = "Student's gender missing at row number " + str(
                            row_no + 1
                        )
                        break
                    else:
                        if check_gender(str(cell.value)):
                            row_data.append(cell.value)
                        else:
                            breaking = error = True
                            error_message = (
                                "Invalid gender input for students data at row number  "
                                + str(row_no + 1)
                                + ", please read the instructions"
                            )
                            break
                elif cell.column_letter == "E":
                    if cell.value == None:
                        breaking = error = True
                        error_message = (
                            "Student's date of birth missing at row number "
                            + str(row_no + 1)
                        )
                        break
                    else:
                        if re.match(
                            "^[0-9]{2}/[0-9]{2}/[0-9]{4}$", str(cell.value)
                        ) and valid_date(str(cell.value)):
                            dob = cell.value[6:] + cell.value[:2] + cell.value[3:5]
                            if valid_dob(dob):
                                row_data.append(cell.value)
                            else:
                                breaking = error = True
                                error_message = (
                                    "Student at row number "
                                    + str(row_no + 1)
                                    + " is not above the age of 5"
                                )
                            break
                        else:
                            breaking = error = True
                            error_message = (
                                "Invalid student's date of birth at row number "
                                + str(row_no + 1)
                            )
                            break
                elif cell.column_letter == "F":
                    if cell.value == None:
                        breaking = error = True
                        error_message = (
                            "Student's pincode missing at row number " + str(row_no + 1)
                        )
                        break
                    else:
                        if not valid_pincode(cell.value):
                            breaking = error = True
                            error_message = (
                                "Invalid student's pincode at row number "
                                + str(row_no + 1)
                            )
                            break
                        else:
                            row_data.append(cell.value)
                elif cell.column_letter == "G":
                    if cell.value == None:
                        breaking = error = True
                        error_message = "Student's state missing at row number " + str(
                            row_no + 1
                        )
                        break
                    else:
                        temp = check_state_city(True, 0, str(cell.value))
                        if temp[0]:
                            row_data.append(cell.value)
                            row_data.append(temp[1])
                        else:
                            breaking = error = True
                            error_message = (
                                "Invalid student's state at row number "
                                + str(row_no + 1)
                            )
                            break
                elif cell.column_letter == "H":
                    if cell.value == None:
                        breaking = error = True
                        error_message = "Student's city missing at row number " + str(
                            row_no + 1
                        )
                        break
                    else:
                        if check_state_city(False, row_data[-1], str(cell.value)):
                            row_data.pop()
                            row_data.append(cell.value)
                        else:
                            breaking = error = True
                            error_message = (
                                "Invalid student's city at row number "
                                + str(row_no + 1)
                            )
                            break
                elif cell.column_letter == "I":
                    if cell.value == None:
                        breaking = error = True
                        error_message = (
                            "Student's address missing at row number " + str(row_no + 1)
                        )
                        break
                    else:
                        row_data.append(cell.value)
                elif cell.column_letter == "J":
                    if cell.value == None:
                        breaking = error = True
                        error_message = (
                            "Student's registration no missing at row number "
                            + str(row_no + 1)
                        )
                        break
                    else:
                        row_data.append(cell.value)
                elif cell.column_letter == "K":
                    if cell.value == None:
                        dob = row_data[-6]
                        if is_adult_func(dob) == "True":
                            row_data.append("ADULT")
                        else:
                            breaking = error = True
                            error_message = (
                                "Student at row number "
                                + str(row_no + 1)
                                + " is not an adult, please provide parentId for the student"
                            )
                            break
                    else:
                        try:
                            parentId = int(cell.value)
                            if parentId <= len(parent_data):
                                row_data.append(cell.value)
                            else:
                                breaking = error = True
                                error_message = (
                                    "Invalid parentId for student data at row number "
                                    + str(row_no + 1)
                                )
                                break
                        except:
                            breaking = error = True
                            error_message = (
                                "Invalid parentId for student data at row number "
                                + str(row_no + 1)
                            )
                            break
            student_data.append(row_data)

        if error == True:
            return render(
                request,
                "teacher/bulkregistration.html",
                {
                    "page_type": "bulk_register",
                    "my_messages": {"error": error_message},
                },
            )

        # for index, row in enumerate(parent_data):
        #     if index == 0:
        #         continue
        #     # creating parent user
        #     skipparent = False
        #     parentData = ParentsInfo.objects.all()
        #     for parent in parentData:
        #         if encryptionHelper.decrypt(parent.email) == encryptionHelper.decrypt(
        #             row[0]
        #         ):
        #             skipparent = True
        #             break

        #     if not skipparent:
        #         password = "".join(
        #             random.choices(string.ascii_lowercase + string.digits, k=8)
        #         )
        #         parentUser = User(username=row[2])
        #         parentUser.set_password(password)
        #         parentUser.save()
        #         parent_group = Group.objects.get(name="Parents")
        #         parentUser.groups.add(parent_group)
        #         parentUser.save()

        #         # getting data from db for foreign keys
        #         city = City.objects.filter(city__icontains=row[9]).first()
        #         state = State.objects.filter(state__icontains=row[10]).first()
        #         education = Education.objects.filter(
        #             education__icontains=row[11]
        #         ).first()
        #         occupation = Occupation.objects.filter(
        #             occupation__icontains=row[12]
        #         ).first()
        #         religion = ReligiousBelief.objects.filter(
        #             religion__icontains=row[13]
        #         ).first()
        #         familyType = FamilyType.objects.filter(
        #             family__icontains=row[14]
        #         ).first()
        #         # creating parent
        #         parent = ParentsInfo(
        #             email=row[0],
        #             name=row[1],
        #             gender=row[3],
        #             age=row[4],
        #             address=row[5],
        #             pincode=row[6],
        #             no_of_family_members=row[7],
        #             children_count=row[8],
        #             city=city,
        #             state=state,
        #             edu=education,
        #             occupation=occupation,
        #             religion=religion,
        #             type_of_family=familyType,
        #             first_password=password,
        #         )
        #         parent.user = parentUser
        #         parent.save()

        # for index, row in enumerate(student_data):
        #     if index == 0:
        #         continue
        #     # creating student user
        #     skipstudent = StudentsInfo.objects.filter(rollno=row[3]).first()
        #     if not skipstudent:
        #         password = "".join(
        #             random.choices(string.ascii_lowercase + string.digits, k=8)
        #         )
        #         studentUser = User(username=row[1])
        #         studentUser.set_password(password)
        #         studentUser.save()
        #         student_group = Group.objects.get(name="Students")
        #         studentUser.groups.add(student_group)
        #         studentUser.save()

        #         # getting data from db for foreign keys
        #         parentData = ParentsInfo.objects.all()
        #         for tempparent in parentData:
        #             if encryptionHelper.decrypt(tempparent.email) == row[7]:
        #                 parent = tempparent

        #         organization = Organization.objects.filter(
        #             name__icontains=row[6]
        #         ).first()
        #         teacher = TeacherInCharge.objects.filter(user=request.user).first()
        #         # creating student
        #         dob = datetime.strptime(row[5], "%Y-%m-%d %H:%M:%S").strftime(
        #             "%Y-%m-%d"
        #         )
        #         student = StudentsInfo(
        #             name=row[0],
        #             address=row[2],
        #             rollno=row[3],
        #             gender=row[4],
        #             dob=dob,
        #             organization=organization,
        #             first_password=password,
        #             teacher=teacher,
        #         )
        #         student.parent = parent
        #         student.user = studentUser
        #         student.save()

        # messages.success(request, "Registration Completed")
        # return redirect("accounts:bulk_register")


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_teacher, login_url="accounts:forbidden")
@password_change_required
def downloadData(request):

    output = io.BytesIO()
    wb = xlsxwriter.Workbook(output)
    parentSheet = wb.add_worksheet("Parents Data")
    studentSheet = wb.add_worksheet("Students Data")

    # student sheet
    row_num = 0
    studentColumns = [
        "Name",
        "Roll No",
        "DOB",
        "Gender",
        "Address",
        "Organization",
        "Parent's Email",
        "Username",
        "Password",
    ]
    for col_num in range(len(studentColumns)):
        studentSheet.write(
            row_num, col_num, studentColumns[col_num]
        )  # at 0 row 0 column

    teacher = TeacherInCharge.objects.filter(user=request.user).first()
    rows = StudentsInfo.objects.filter(teacher=teacher).values_list(
        "name",
        "rollno",
        "dob",
        "gender",
        "address",
        "organization",
        "parent",
        "user",
        "first_password",
        "password_changed",
    )
    parentEmail = []
    for row in rows:
        row_num += 1
        for col_num in range(len(studentColumns)):
            if col_num == 0:
                studentSheet.write(
                    row_num, col_num, encryptionHelper.decrypt(row[col_num])
                )

            elif col_num == 2:
                studentSheet.write(row_num, col_num, row[col_num].strftime("%d/%b/%Y"))

            elif col_num == 5:
                organization = Organization.objects.get(pk=row[col_num])
                studentSheet.write(row_num, col_num, organization.name)

            elif col_num == 6:
                parent = ParentsInfo.objects.get(pk=row[col_num])
                email = encryptionHelper.decrypt(parent.email)
                if email not in parentEmail:
                    parentEmail.append(email)
                studentSheet.write(row_num, col_num, email)

            elif col_num == 7:
                user = User.objects.get(pk=row[col_num])
                studentSheet.write(row_num, col_num, user.username)

            elif col_num == 8:
                msg = row[col_num]
                if row[col_num + 1]:
                    msg = "Already Changed"
                studentSheet.write(row_num, col_num, msg)

            elif col_num == 9:
                continue

            else:
                studentSheet.write(row_num, col_num, row[col_num])

    # Sheet header, first row
    row_num = 0

    # parent sheet
    parentColumns = [
        "Name",
        "Email",
        "Age",
        "Gender",
        "Address",
        "City",
        "State",
        "Pincode",
        "Education",
        "Occupation",
        "Religion",
        "Family count",
        "Children count",
        "Family Type",
        "Username",
        "Password",
    ]
    for col_num in range(len(parentColumns)):
        parentSheet.write(row_num, col_num, parentColumns[col_num])  # at 0 row 0 column

    rows = ParentsInfo.objects.all().values_list(
        "name",
        "email",
        "age",
        "gender",
        "address",
        "city",
        "state",
        "pincode",
        "edu",
        "occupation",
        "religion",
        "no_of_family_members",
        "children_count",
        "type_of_family",
        "user",
        "first_password",
        "password_changed",
    )
    for row in rows:
        if encryptionHelper.decrypt(row[1]) in parentEmail:
            row_num += 1
            for col_num in range(len(parentColumns)):

                if col_num == 0 or col_num == 1:
                    parentSheet.write(
                        row_num, col_num, encryptionHelper.decrypt(row[col_num])
                    )

                elif col_num == 5:
                    city = City.objects.get(pk=row[col_num])
                    parentSheet.write(row_num, col_num, city.city)

                elif col_num == 6:
                    state = State.objects.get(pk=row[col_num])
                    parentSheet.write(row_num, col_num, state.state)

                elif col_num == 8:
                    education = Education.objects.get(pk=row[col_num])
                    parentSheet.write(row_num, col_num, education.education)

                elif col_num == 9:
                    occupation = Occupation.objects.get(pk=row[col_num])
                    parentSheet.write(row_num, col_num, occupation.occupation)

                elif col_num == 10:
                    religion = ReligiousBelief.objects.get(pk=row[col_num])
                    parentSheet.write(row_num, col_num, religion.religion)

                elif col_num == 13:
                    familyType = FamilyType.objects.get(pk=row[col_num])
                    parentSheet.write(row_num, col_num, familyType.family)

                elif col_num == 14:
                    user = User.objects.get(pk=row[col_num])
                    parentSheet.write(row_num, col_num, user.username)

                elif col_num == 15:
                    msg = row[col_num]
                    if row[col_num + 1]:
                        msg = "Already Changed"
                    parentSheet.write(row_num, col_num, msg)

                elif col_num == 16:
                    continue

                else:
                    parentSheet.write(row_num, col_num, row[col_num])

    wb.close()
    # construct response
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response[
        "Content-Disposition"
    ] = "attachment; filename=Parent and Student list.xlsx"
    return response


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_student, login_url="accounts:forbidden")
@password_change_required
def student_dashboard(request):
    return render(
        request,
        "student/student_dashboard.html",
        {"page_type": "student_dashboard"},
    )


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_teacher, login_url="accounts:forbidden")
@password_change_required
def viewSessionForms(request, id):
    teacher = TeacherInCharge.objects.get(user=request.user)
    teachersession = Session.objects.filter(id=id).first()
    total_students = StudentsInfo.objects.filter(
        session=teachersession, teacher=teacher
    )
    results = []
    closed_sessions = FormDetails.objects.filter(
        teacher=teacher, session=teachersession, open=False
    )

    for session in closed_sessions:
        temp_list = [session.form, session.start_timestamp, session.end_timestamp]
        if session.pre:
            temp_list.append("Pre Test")
        else:
            temp_list.append("Post Test")
        count = 0
        for student in total_students:
            if ModuleOne.objects.filter(
                student=student,
                submission_timestamp__gte=session.start_timestamp,
                submission_timestamp__lte=session.end_timestamp,
            ).exists():
                draftForm = ModuleOne.objects.filter(
                    student=student,
                    submission_timestamp__gte=session.start_timestamp,
                    submission_timestamp__lte=session.end_timestamp,
                ).first()
                if not draftForm.draft:
                    count += 1
            elif Activity.objects.filter(
                student=student,
                submission_timestamp__gte=session.start_timestamp,
                submission_timestamp__lte=session.end_timestamp,
            ).exists():
                draftForm = Activity.objects.filter(
                    student=student,
                    submission_timestamp__gte=session.start_timestamp,
                    submission_timestamp__lte=session.end_timestamp,
                ).first()
                if not draftForm.draft:
                    count += 1

        temp_list.append(count)
        temp_list.append(len(total_students))
        temp_list.append(session.id)
        results.append(temp_list)
    open_sessions = FormDetails.objects.filter(
        teacher=teacher, session=teachersession, open=True
    )
    results2 = []
    for session in open_sessions:
        temp_list = [session.form, session.start_timestamp]
        if session.pre:
            temp_list.append("Pre Test")
        else:
            temp_list.append("Post Test")
        count = 0
        for student in total_students:
            if ModuleOne.objects.filter(
                student=student, submission_timestamp__gte=session.start_timestamp
            ).exists():
                draftForm = ModuleOne.objects.filter(
                    student=student, submission_timestamp__gte=session.start_timestamp
                ).first()
                if not draftForm.draft:
                    count += 1
            elif Activity.objects.filter(
                student=student, submission_timestamp__gte=session.start_timestamp
            ).exists():
                draftForm = Activity.objects.filter(
                    student=student, submission_timestamp__gte=session.start_timestamp
                ).first()
                if not draftForm.draft:
                    count += 1

        temp_list.append(count)
        temp_list.append(len(total_students))
        temp_list.append(session.id)
        results2.append(temp_list)

    return render(
        request,
        "teacher/view_session_forms.html",
        {
            "results": results,
            "results2": results2,
            "page_type": "view_session_forms",
            "session": teachersession,
        },
    )


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_coordinator, login_url="accounts:forbidden")
@password_change_required
def coordinator_dashboard(request):
    teachers = (
        CoordinatorInCharge.objects.filter(user=request.user)
        .first()
        .teacherincharge_set.all()
    )
    for teacher in teachers:
        teacher.name = encryptionHelper.decrypt(teacher.name)
    return render(
        request,
        "coordinator/coordinator_dashboard.html",
        {"teachers": teachers, "page_type": "coordinator_dashboard"},
    )


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_supercoordinator, login_url="accounts:forbidden")
def supercoordinator_dashboard(request):
    organizations = Organization.objects.all()

    return render(
        request,
        "supercoordinator/supercoordinator_dashboard.html",
        {"organizations": organizations, "page_type": "supercoordinator_dashboard"},
    )


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_supercoordinator, login_url="accounts:forbidden")
def viewCoordinators(request, id):
    organization = Organization.objects.filter(id=id).first()
    coordinators = organization.coordinatorincharge_set.all()
    for coordinator in coordinators:
        coordinator.name = encryptionHelper.decrypt(coordinator.name)
        coordinator.mobile_no = encryptionHelper.decrypt(coordinator.mobile_no)
        coordinator.email = encryptionHelper.decrypt(coordinator.email)
    return render(
        request,
        "supercoordinator/view_coordinators.html",
        {
            "coordinators": coordinators,
            "organization": organization,
            "page_type": "view_coordinators",
            "org_id": id,
        },
    )


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_supercoordinator, login_url="accounts:forbidden")
def allCoordinators(request):
    coordinators = CoordinatorInCharge.objects.all()
    for coordinator in coordinators:
        coordinator.name = encryptionHelper.decrypt(coordinator.name)
        coordinator.mobile_no = encryptionHelper.decrypt(coordinator.mobile_no)
        coordinator.email = encryptionHelper.decrypt(coordinator.email)
    return render(
        request,
        "supercoordinator/all_coordinators.html",
        {"coordinators": coordinators, "page_type": "all_coordinators"},
    )


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_coordinator, login_url="accounts:forbidden")
@password_change_required
def allSessions(request):
    sessions = (
        CoordinatorInCharge.objects.filter(user=request.user).first().session_set.all()
    )
    open_sessions = []
    close_sessions = []
    open = False
    close = False
    for session in sessions:
        if session.end_date == None:
            open_sessions.append(session)
            open = True
        else:
            close_sessions.append(session)
            close = True
    return render(
        request,
        "coordinator/all_sessions.html",
        {
            "open_sessions": open_sessions,
            "close_sessions": close_sessions,
            "open": open,
            "close": close,
            "page_type": "all_sessions",
        },
    )


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_teacher, login_url="accounts:forbidden")
@password_change_required
def teacherAllSessions(request):
    teacher = TeacherInCharge.objects.filter(user=request.user).first()
    objects = Teacher_Session.objects.filter(teacher=teacher)
    sessions = []
    for object in objects:
        sessions.append(object.session)
    open_sessions = []
    close_sessions = []
    open = False
    close = False
    for session in sessions:
        if session.end_date == None:
            open_sessions.append(session)
            open = True
        else:
            close_sessions.append(session)
            close = True
    return render(
        request,
        "teacher/teacher_all_sessions.html",
        {
            "open_sessions": open_sessions,
            "close_sessions": close_sessions,
            "open": open,
            "close": close,
            "page_type": "teacher_all_sessions",
        },
    )


def createTempDict(postData):
    temp = {}
    for key in postData:
        if key == "source_fruits_vegetables" or key == "grow_own_food":
            temp[key] = postData.getlist(key)
        else:
            temp[key] = postData[key]
    del temp["csrfmiddlewaretoken"]
    return temp


def creatingOrUpdatingDrafts(temp, user, formName):
    student = StudentsInfo.objects.get(user=user)
    startdate = FormDetails.objects.get(
        form=Form.objects.get(name=formName), teacher=student.teacher, open=True
    ).start_timestamp
    if ModuleOne.objects.filter(
        student=student, submission_timestamp__gte=startdate
    ).exists():
        draftForm = ModuleOne.objects.get(
            student=StudentsInfo.objects.get(user=user),
            submission_timestamp__gte=startdate,
        )
        if draftForm.draft:
            # updating drafts
            for name in ModuleOne._meta.get_fields():
                name = name.column
                if name == "id" or name == "student_id" or name == "draft":
                    continue
                if name in temp:
                    setattr(draftForm, name, temp[name])
                else:
                    setattr(draftForm, name, getattr(draftForm, name) or None)

            draftForm.submission_timestamp = datetime.now()
            draftForm.save()
            return True
    else:
        return False


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_parent or is_student, login_url="accounts:forbidden")
@password_change_required
def draft(request):
    if "parent_dashboard" in request.META.get("HTTP_REFERER").split("/"):
        module = request.META.get("HTTP_REFERER").split("/")[-1]
        id = request.META.get("HTTP_REFERER").split("/")[-2]
        user = StudentsInfo.objects.get(pk=id).user
    else:
        module = request.META.get("HTTP_REFERER").split("/")[-2]
        user = request.user

    # 1st Page
    if module == "moduleOne":
        # for removing csrf field
        temp = createTempDict(request.POST)
        # checking if draft exists
        if not creatingOrUpdatingDrafts(temp, user, "moduleOne"):
            # creating new record
            form = ModuleOne(**temp)
            form.student = StudentsInfo.objects.get(user=user)
            form.draft = True
            formType = getFormType(
                "moduleOne", StudentsInfo.objects.get(user=user).teacher
            )
            form.pre = 1 if formType == "PreTest" else 0
            form.submission_timestamp = datetime.now()
            form.save()
    # 2nd and 3rd Page
    elif module == "moduleOne-2" or module == "moduleOne-3":
        temp = createTempDict(request.POST)
        creatingOrUpdatingDrafts(temp, user, "moduleOne")
    return redirect(request.META.get("HTTP_REFERER"))


def getFormType(moduleType, teacher):
    module = Form.objects.get(name=moduleType)
    formType = FormDetails.objects.filter(
        form=module, open=True, teacher=teacher
    ).first()
    if formType.pre:
        return "PreTest"
    else:
        return "PostTest"


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_parent or is_student, login_url="accounts:forbidden")
@password_change_required
@isActive("moduleOne", "student")
def moduleOne(request, user=None):
    if request.method == "GET":
        if user == None:
            user = request.user

        student = StudentsInfo.objects.get(user=user)
        startdate = FormDetails.objects.get(
            form=Form.objects.get(name="moduleOne"), teacher=student.teacher, open=True
        ).start_timestamp
        if ModuleOne.objects.filter(
            student=student, submission_timestamp__gte=startdate
        ).exists():
            draftForm = ModuleOne.objects.get(
                student=StudentsInfo.objects.get(user=user),
                submission_timestamp__gte=startdate,
            )
            if draftForm.draft:
                mod = ModuleOneForm()
                temp = {}
                for name in ModuleOne._meta.get_fields():
                    name = name.column
                    if name in mod.fields:
                        if (
                            name == "source_fruits_vegetables"
                            or name == "grow_own_food"
                        ):
                            temp[name] = ast.literal_eval(
                                getattr(draftForm, name) or "[]"
                            )
                        else:
                            temp[name] = getattr(draftForm, name)

                form = ModuleOneForm(temp)
                formPre = getFormType("moduleOne", student.teacher)
                return render(
                    request,
                    "moduleOne/module_one.html",
                    {
                        "form": form,
                        "formPre": formPre,
                        "page_type": "student_module_one",
                    },
                )
            else:
                return redirect("../already_filled")

        # new form
        else:
            form = ModuleOneForm()
            formPre = getFormType("moduleOne", student.teacher)
            return render(
                request,
                "moduleOne/module_one.html",
                {"form": form, "formPre": formPre, "page_type": "student_module_one"},
            )

    # POST
    else:

        flag = False
        if user == None:
            flag = True
            user = request.user
        form = ModuleOneForm(request.POST)

        if form.is_valid():
            temp = createTempDict(request.POST)

            if not creatingOrUpdatingDrafts(temp, user, "moduleOne"):
                # creating new record
                form = ModuleOne(**temp)
                form.student = StudentsInfo.objects.get(user=user)
                form.draft = True
                formType = getFormType(
                    "moduleOne", StudentsInfo.objects.get(user=user).teacher
                )
                form.pre = 1 if formType == "PreTest" else 0
                form.submission_timestamp = datetime.now()
                form.save()

            if flag:
                return redirect("accounts:module_one_2")
            else:
                return redirect(
                    "accounts:parentsModuleOne2",
                    id=StudentsInfo.objects.get(user=user).id,
                )

        else:
            formPre = getFormType(
                "moduleOne", StudentsInfo.objects.get(user=user).teacher
            )
            return render(
                request,
                "moduleOne/module_one.html",
                {"form": form, "formPre": formPre, "page_type": "student_module_one"},
            )


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_parent or is_student, login_url="accounts:forbidden")
@password_change_required
@isActive("moduleOne", "student")
def moduleOne2(request, user=None):
    if request.method == "GET":
        if user == None:
            user = request.user

        student = StudentsInfo.objects.get(user=user)
        startdate = FormDetails.objects.get(
            form=Form.objects.get(name="moduleOne"), teacher=student.teacher, open=True
        ).start_timestamp
        if ModuleOne.objects.filter(
            student=student, submission_timestamp__gte=startdate
        ).exists():
            draftForm = ModuleOne.objects.get(
                student=StudentsInfo.objects.get(user=user),
                submission_timestamp__gte=startdate,
            )
            if draftForm.draft:
                mod = ModuleOneForm2()
                temp = {}
                for name in ModuleOne._meta.get_fields():
                    name = name.column
                    if name in mod.fields:
                        temp[name] = getattr(draftForm, name) or None

                form = ModuleOneForm2(temp)
                formPre = getFormType("moduleOne", student.teacher)
                return render(
                    request,
                    "moduleOne/module_one2.html",
                    {
                        "form": form,
                        "formPre": formPre,
                        "page_type": "student_module_one",
                    },
                )
            else:
                return redirect("../already_filled")

        # new form
        else:
            form = ModuleOneForm2()
            formPre = getFormType("moduleOne", student.teacher)
            return render(
                request,
                "moduleOne/module_one2.html",
                {"form": form, "formPre": formPre, "page_type": "student_module_one"},
            )
    # POST
    else:
        flag = False
        if user == None:
            flag = True
            user = request.user
        form = ModuleOneForm2(request.POST)

        if form.is_valid():
            temp = createTempDict(request.POST)
            creatingOrUpdatingDrafts(temp, user, "moduleOne")

            if flag:
                return redirect("accounts:module_one_3")
            else:
                return redirect(
                    "accounts:parentsModuleOne3",
                    id=StudentsInfo.objects.get(user=user).id,
                )
        else:
            formPre = getFormType(
                "moduleOne", StudentsInfo.objects.get(user=user).teacher
            )
            return render(
                request,
                "moduleOne/module_one2.html",
                {"form": form, "formPre": "formPre", "page_type": "student_module_one"},
            )


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_parent or is_student, login_url="accounts:forbidden")
@password_change_required
@isActive("moduleOne", "student")
def moduleOne3(request, user=None):
    if request.method == "GET":
        if user == None:
            user = request.user

        student = StudentsInfo.objects.get(user=user)
        startdate = FormDetails.objects.get(
            form=Form.objects.get(name="moduleOne"), teacher=student.teacher, open=True
        ).start_timestamp
        if ModuleOne.objects.filter(
            student=student, submission_timestamp__gte=startdate
        ).exists():
            draftForm = ModuleOne.objects.get(
                student=StudentsInfo.objects.get(user=user),
                submission_timestamp__gte=startdate,
            )
            if draftForm.draft:
                mod = ModuleOneForm3()
                temp = {}
                for name in ModuleOne._meta.get_fields():
                    name = name.column
                    if name in mod.fields:
                        temp[name] = getattr(draftForm, name) or None
                form = ModuleOneForm3(temp)
                formPre = getFormType("moduleOne", student.teacher)
                return render(
                    request,
                    "moduleOne/module_one3.html",
                    {
                        "form": form,
                        "formPre": formPre,
                        "page_type": "student_module_one",
                    },
                )
            else:
                return redirect("../already_filled")
        # new form
        else:
            form = ModuleOneForm3()
            formPre = getFormType("moduleOne", student.teacher)
            return render(
                request,
                "moduleOne/module_one3.html",
                {"form": form, "formPre": formPre, "page_type": "student_module_one"},
            )
    # POST
    else:
        flag = False
        if user == None:
            flag = True
            user = request.user
        form = ModuleOneForm3(request.POST)

        # valid form
        if form.is_valid():
            temp = createTempDict(request.POST)
            student = StudentsInfo.objects.get(user=user)
            startdate = FormDetails.objects.get(
                form=Form.objects.get(name="moduleOne"),
                teacher=student.teacher,
                open=True,
            ).start_timestamp
            draftForm = ModuleOne.objects.get(
                student=StudentsInfo.objects.get(user=user),
                submission_timestamp__gte=startdate,
            )
            if draftForm.draft:
                for name in ModuleOne._meta.get_fields():
                    name = name.column
                    if name == "id" or name == "student_id" or name == "draft":
                        continue
                    elif name == "source_fruits_vegetables" or name == "grow_own_food":
                        my_list = "; ".join(ast.literal_eval(getattr(draftForm, name)))
                        setattr(draftForm, name, my_list)
                    elif name in temp:
                        setattr(draftForm, name, temp[name])

                draftForm.draft = False
                draftForm.submission_timestamp = datetime.now()
                draftForm.save()
                if flag:
                    return redirect("accounts:student_dashboard")
                else:
                    return redirect("accounts:parent_dashboard")
        # invalid form
        else:
            formPre = getFormType(
                "moduleOne", StudentsInfo.objects.get(user=user).teacher
            )
            return render(
                request,
                "moduleOne/module_one3.html",
                {"form": form, "formPre": formPre, "page_type": "student_module_one"},
            )


def forbidden(request):
    login = False
    if request.user.get_username() != "":
        login = True
    return render(request, "other/forbidden.html", {"login": login})


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_parent, login_url="accounts:forbidden")
@password_change_required
def showStudent(request, id):
    student = StudentsInfo.objects.get(pk=id)
    return render(request, "parent/student_modules.html", {"student": student})


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_parent, login_url="accounts:forbidden")
@password_change_required
@isActive("moduleOne", "parent")
def parentModuleOne(request, id):
    user = StudentsInfo.objects.get(pk=id).user
    return moduleOne(request, user)


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_parent, login_url="accounts:forbidden")
@password_change_required
@isActive("moduleOne", "parent")
def parentModuleOne2(request, id):
    user = StudentsInfo.objects.get(pk=id).user
    return moduleOne2(request, user)


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_parent, login_url="accounts:forbidden")
@password_change_required
@isActive("moduleOne", "parent")
def parentModuleOne3(request, id):
    user = StudentsInfo.objects.get(pk=id).user
    return moduleOne3(request, user)


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_parent or is_student, login_url="accounts:forbidden")
@password_change_required
def previous(request):
    link = request.META.get("HTTP_REFERER").split("/")
    if "parent_dashboard" in link:
        if link[-1] == "moduleOne-2":
            link[-1] = "moduleOne"
        elif link[-1] == "moduleOne-3":
            link[-1] = "moduleOne-2"
    else:
        if link[-2] == "moduleOne-2":
            link[-2] = "moduleOne"
        elif link[-2] == "moduleOne-3":
            link[-2] = "moduleOne-2"
    newLink = "/".join(link)
    return redirect(newLink)


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_teacher, login_url="accounts:forbidden")
@password_change_required
def manageForms(request, id):
    if request.method == "GET":
        activity = {}
        moduleOne = {}
        moduleTwo = {}
        moduleThree = {}
        teacher = TeacherInCharge.objects.get(user=request.user)
        session = Session.objects.filter(id=id).first()
        form = Form.objects.get(name="moduleOne")
        if FormDetails.objects.filter(
            form=form, teacher=teacher, session=session
        ).exists():
            form = (
                FormDetails.objects.filter(
                    form=form, teacher=teacher, session=session, open=True
                )
                .order_by("-start_timestamp")
                .first()
            )
            if form:
                if form.pre:
                    moduleOne["pre"] = True
                    moduleOne["post"] = False

                else:
                    moduleOne["post"] = True
                    moduleOne["pre"] = False

        activity_form = Form.objects.get(name="activity")
        if FormDetails.objects.filter(
            form=activity_form, teacher=teacher, session=session
        ).exists():
            form2 = (
                FormDetails.objects.filter(
                    form=activity_form, teacher=teacher, session=session, open=True
                )
                .order_by("-start_timestamp")
                .first()
            )
            print(form2)
            if form2:
                print("pre")
                if form2.pre:
                    activity["pre"] = True
                    activity["post"] = False

                else:
                    activity["post"] = True
                    activity["pre"] = False

        return render(
            request,
            "teacher/manage_forms_teacher.html",
            {
                "activity": activity,
                "moduleOne": moduleOne,
                "moduleTwo": moduleTwo,
                "moduleThree": moduleThree,
                "page_type": "manage_forms",
            },
        )
    else:

        if "moduleOne" in request.POST:
            module_one_pre = request.POST.get("module_one_pre", False)
            module_one_post = request.POST.get("module_one_post", False)
            if module_one_pre == "on" and module_one_post == "on":
                messages.error(request, "Cannot select both PreTest and PostTest")
                return redirect("accounts:manage_forms")

            form = Form.objects.get(name="moduleOne")
            teacher = TeacherInCharge.objects.get(user=request.user)
            session = Session.objects.filter(id=id).first()
            if module_one_pre == "on":
                if not FormDetails.objects.filter(
                    form=form, teacher=teacher, session=session, pre=True, open=True
                ).exists():
                    update = FormDetails(
                        form=form,
                        teacher=teacher,
                        session=session,
                        pre=True,
                        open=True,
                        start_timestamp=datetime.now(),
                    )
                    update.save()
            else:
                if FormDetails.objects.filter(
                    form=form, teacher=teacher, session=session, pre=True, open=True
                ).exists():
                    update = FormDetails.objects.filter(
                        form=form, teacher=teacher, session=session, pre=True, open=True
                    ).first()
                    update.open = False
                    update.end_timestamp = datetime.now()
                    session = Session.objects.filter(id=id).first()
                    teacher = TeacherInCharge.objects.filter(user=request.user).first()
                    total_students = StudentsInfo.objects.filter(
                        session=session, teacher=teacher
                    )
                    for student in total_students:
                        if ModuleOne.objects.filter(
                            student=student,
                            submission_timestamp__gte=update.start_timestamp,
                            submission_timestamp__lte=update.end_timestamp,
                            draft=True,
                            pre=True,
                        ).exists():
                            draftForm = ModuleOne.objects.filter(
                                student=student,
                                submission_timestamp__gte=update.start_timestamp,
                                submission_timestamp__lte=update.end_timestamp,
                                draft=True,
                                pre=True,
                            ).first()
                            draftForm.delete()
                    update.save()

            if module_one_post == "on":
                if not FormDetails.objects.filter(
                    form=form, teacher=teacher, session=session, pre=False, open=True
                ).exists():
                    update = FormDetails(
                        form=form,
                        teacher=teacher,
                        session=session,
                        pre=False,
                        open=True,
                        start_timestamp=datetime.now(),
                    )
                    update.save()
            else:
                if FormDetails.objects.filter(
                    form=form, teacher=teacher, session=session, pre=False, open=True
                ).exists():
                    update = FormDetails.objects.filter(
                        form=form,
                        teacher=teacher,
                        session=session,
                        pre=False,
                        open=True,
                    ).first()
                    update.open = False
                    update.end_timestamp = datetime.now()
                    teacher = TeacherInCharge.objects.get(user=request.user)
                    session = Session.objects.filter(id=id).first()
                    total_students = StudentsInfo.objects.filter(
                        session=session, teacher=teacher
                    )
                    for student in total_students:
                        if ModuleOne.objects.filter(
                            student=student,
                            submission_timestamp__gte=update.start_timestamp,
                            submission_timestamp__lte=update.end_timestamp,
                            draft=True,
                            pre=False,
                        ).exists():
                            draftForm = ModuleOne.objects.filter(
                                student=student,
                                submission_timestamp__gte=update.start_timestamp,
                                submission_timestamp__lte=update.end_timestamp,
                                draft=True,
                                pre=False,
                            ).first()
                            draftForm.delete()
                    update.save()

        elif "activity" in request.POST:
            activity_pre = request.POST.get("activity_pre", False)
            activity_post = request.POST.get("activity_post", False)
            if activity_pre == "on" and activity_post == "on":
                messages.error(request, "Cannot select both PreTest and PostTest")
                return redirect("accounts:manage_forms")

            form = Form.objects.get(name="activity")
            teacher = TeacherInCharge.objects.get(user=request.user)
            session = Session.objects.filter(id=id).first()
            total_students = StudentsInfo.objects.filter(
                session=session, teacher=teacher
            )

            if activity_pre == "on":
                if not FormDetails.objects.filter(
                    form=form, teacher=teacher, session=session, pre=True, open=True
                ).exists():
                    update = FormDetails(
                        form=form,
                        teacher=teacher,
                        session=session,
                        pre=True,
                        open=True,
                        start_timestamp=datetime.now(),
                    )
                    update.save()
            else:
                if FormDetails.objects.filter(
                    form=form, teacher=teacher, session=session, pre=True, open=True
                ).exists():
                    update = FormDetails.objects.filter(
                        form=form, teacher=teacher, session=session, pre=True, open=True
                    ).first()
                    update.open = False
                    update.end_timestamp = datetime.now()
                    teacher = TeacherInCharge.objects.get(user=request.user)
                    session = Session.objects.filter(id=id).first()
                    total_students = StudentsInfo.objects.filter(
                        session=session, teacher=teacher
                    )
                    for student in total_students:
                        if Activity.objects.filter(
                            student=student,
                            submission_timestamp__gte=update.start_timestamp,
                            submission_timestamp__lte=update.end_timestamp,
                            draft=True,
                            pre=True,
                        ).exists():
                            draftForm = Activity.objects.filter(
                                student=student,
                                submission_timestamp__gte=update.start_timestamp,
                                submission_timestamp__lte=update.end_timestamp,
                                draft=True,
                                pre=True,
                            ).first()
                            draftForm.delete()
                    update.save()

            if activity_post == "on":
                if not FormDetails.objects.filter(
                    form=form, teacher=teacher, session=session, pre=False, open=True
                ).exists():
                    update = FormDetails(
                        form=form,
                        teacher=teacher,
                        session=session,
                        pre=False,
                        open=True,
                        start_timestamp=datetime.now(),
                    )
                    update.save()
            else:
                if FormDetails.objects.filter(
                    form=form, teacher=teacher, session=session, pre=False, open=True
                ).exists():
                    update = FormDetails.objects.filter(
                        form=form,
                        teacher=teacher,
                        session=session,
                        pre=False,
                        open=True,
                    ).first()
                    update.open = False
                    update.end_timestamp = datetime.now()
                    teacher = TeacherInCharge.objects.get(user=request.user)
                    session = Session.objects.filter(id=id).first()
                    total_students = StudentsInfo.objects.filter(
                        session=session, teacher=teacher
                    )
                    for student in total_students:
                        if Activity.objects.filter(
                            student=student,
                            submission_timestamp__gte=update.start_timestamp,
                            submission_timestamp__lte=update.end_timestamp,
                            draft=True,
                            pre=False,
                        ).exists():
                            draftForm = Activity.objects.filter(
                                student=student,
                                submission_timestamp__gte=update.start_timestamp,
                                submission_timestamp__lte=update.end_timestamp,
                                draft=True,
                                pre=False,
                            ).first()
                            draftForm.delete()
                    update.save()

        module_two_pre = request.POST.get("module_two_pre", False)
        module_two_post = request.POST.get("module_two_post", False)
        module_three_pre = request.POST.get("module_three_pre", False)
        module_three_post = request.POST.get("module_three_post", False)

        return redirect("accounts:manage_forms", id)


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_teacher, login_url="accounts:forbidden")
@password_change_required
def getFormDetails(request, id):
    form = FormDetails.objects.get(pk=id)
    teacher = form.teacher
    total_students = teacher.studentsinfo_set.all()

    filled_students = []
    not_filled_students = []
    if not form.open:
        form_open = False
        temp_list = [form.form, form.start_timestamp, form.end_timestamp]
    else:
        form_open = True
        temp_list = [form.form, form.start_timestamp]

    if form.pre:
        temp_list.append("Pre Test")
    else:
        temp_list.append("Post Test")

    if form.form.name == "moduleOne":
        for student in total_students:
            temp = []
            if form.open:
                if ModuleOne.objects.filter(
                    student=student, submission_timestamp__gte=form.start_timestamp
                ).exists():
                    draftForm = ModuleOne.objects.filter(
                        student=student, submission_timestamp__gte=form.start_timestamp
                    ).first()

                    if draftForm.draft:
                        temp.append(encryptionHelper.decrypt(student.name))
                        temp.append("-")
                        not_filled_students.append(temp)
                    else:
                        temp.append(encryptionHelper.decrypt(student.name))
                        temp.append(draftForm.submission_timestamp)
                        filled_students.append(temp)
                else:
                    temp.append(encryptionHelper.decrypt(student.name))
                    temp.append("-")
                    not_filled_students.append(temp)

            else:
                if ModuleOne.objects.filter(
                    student=student,
                    submission_timestamp__gte=form.start_timestamp,
                    submission_timestamp__lte=form.end_timestamp,
                ).exists():
                    submitted_form = ModuleOne.objects.filter(
                        student=student, submission_timestamp__gte=form.start_timestamp
                    ).first()
                    temp.append(encryptionHelper.decrypt(student.name))
                    temp.append(submitted_form.submission_timestamp)
                    filled_students.append(temp)
                else:
                    temp.append(encryptionHelper.decrypt(student.name))
                    temp.append("-")
                    not_filled_students.append(temp)

    elif form.form.name == "activity":
        for student in total_students:
            temp = []
            if form.open:
                if Activity.objects.filter(
                    student=student, submission_timestamp__gte=form.start_timestamp
                ).exists():
                    draftForm = Activity.objects.filter(
                        student=student, submission_timestamp__gte=form.start_timestamp
                    ).first()

                    if draftForm.draft:
                        temp.append(encryptionHelper.decrypt(student.name))
                        temp.append("-")
                        not_filled_students.append(temp)
                    else:
                        temp.append(encryptionHelper.decrypt(student.name))
                        temp.append(draftForm.submission_timestamp)
                        filled_students.append(temp)
                else:
                    temp.append(encryptionHelper.decrypt(student.name))
                    temp.append("-")
                    not_filled_students.append(temp)

            else:
                if Activity.objects.filter(
                    student=student,
                    submission_timestamp__gte=form.start_timestamp,
                    submission_timestamp__lte=form.end_timestamp,
                ).exists():
                    submitted_form = Activity.objects.filter(
                        student=student, submission_timestamp__gte=form.start_timestamp
                    ).first()
                    temp.append(encryptionHelper.decrypt(student.name))
                    temp.append(submitted_form.submission_timestamp)
                    filled_students.append(temp)
                else:
                    temp.append(encryptionHelper.decrypt(student.name))
                    temp.append("-")
                    not_filled_students.append(temp)

    return render(
        request,
        "teacher/teacher_dashboard_getDetails.html",
        {
            "result": temp_list,
            "filled_students": filled_students,
            "not_filled_students": not_filled_students,
            "open": form_open,
        },
    )


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_supercoordinator, login_url="accounts:forbidden")
def supercoordinator_reset_password(request):
    if request.method == "GET":
        form = SuperCoordPasswordReset()
        return render(
            request,
            "supercoordinator/supercoordinator_reset_password.html",
            {"form": form, "page_type": "reset_password"},
        )
    else:
        form = SuperCoordPasswordReset(request.POST)
        if form.is_valid():
            username = form.cleaned_data["username"]
            if User.objects.filter(username=username).exists():
                user = User.objects.get(username=username)
                if is_coordinator(user):
                    password = random_password_generator()
                    user_coord = CoordinatorInCharge.objects.get(user=user)
                    user_coord.first_password = password
                    user_coord.password_changed = False
                    user_coord.save()
                    user.set_password(password)
                    user.save()
                    return render(
                        request,
                        "supercoordinator/supercoordinator_reset_password.html",
                        {
                            "form": form,
                            "page_type": "reset_password",
                            "my_messages": {
                                "success": "Password reset successfull. Download login credentions from below."
                            },
                        },
                    )
                else:
                    return render(
                        request,
                        "supercoordinator/supercoordinator_reset_password.html",
                        {
                            "form": form,
                            "page_type": "reset_password",
                            "my_messages": {
                                "error": "Provided user is not a Coordinator."
                            },
                        },
                    )
            else:
                return render(
                    request,
                    "supercoordinator/supercoordinator_reset_password.html",
                    {
                        "form": form,
                        "page_type": "reset_password",
                        "my_messages": {"error": "Invalid username."},
                    },
                )
        else:
            return render(
                request,
                "supercoordinator/supercoordinator_reset_password.html",
                {"form": form, "page_type": "reset_password"},
            )


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_supercoordinator, login_url="accounts:forbidden")
def supercoordinator_reset_password_download(request):
    output = io.BytesIO()
    wb = xlsxwriter.Workbook(output)
    bold = wb.add_format({"bold": True})
    credentials = wb.add_worksheet("Credentials")
    columns = [
        "Username",
        "Password",
    ]
    for col_num in range(len(columns)):
        credentials.write(0, col_num, columns[col_num], bold)
    coord = CoordinatorInCharge.objects.filter(password_changed=False)
    for row_no, x in enumerate(coord):
        credentials.write(row_no + 1, 0, x.user.username)
        credentials.write(row_no + 1, 1, x.first_password)
    wb.close()
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = "attachment; filename=Login Credentials.xlsx"
    return response


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_coordinator, login_url="accounts:forbidden")
@password_change_required
def coordinator_reset_password(request):
    if request.method == "GET":
        form = CoordPasswordReset()
        return render(
            request,
            "coordinator/coordinator_reset_password.html",
            {"form": form, "page_type": "reset_password"},
        )
    else:
        form = CoordPasswordReset(request.POST)
        organization = CoordinatorInCharge.object.get(user=request.user).organization
        if form.is_valid():
            username = form.cleaned_data["username"]
            if User.objects.filter(username=username).exists():
                user = User.objects.get(username=username)
                if is_teacher(user):
                    user_teacher = TeacherInCharge.objects.get(user=user)
                    if user_teacher.organization == organization:
                        password = random_password_generator()
                        user_teacher.first_password = password
                        user_teacher.password_changed = False
                        user_teacher.save()
                        user.set_password(password)
                        user.save()
                        return render(
                            request,
                            "coordinator/coordinator_reset_password.html",
                            {
                                "form": form,
                                "page_type": "reset_password",
                                "my_messages": {
                                    "success": "Password reset successfull. Download login credentions from below."
                                },
                            },
                        )
                    else:
                        return render(
                            request,
                            "coordinator/coordinator_reset_password.html",
                            {
                                "form": form,
                                "page_type": "reset_password",
                                "my_messages": {
                                    "error": "Teacher does not belong to your organization."
                                },
                            },
                        )
                elif is_parent(user):
                    user_parent = ParentsInfo.objects.get(user=user)
                    if user_parent.organization == organization:
                        password = random_password_generator()
                        user_parent.first_password = password
                        user_parent.password_changed = False
                        user_parent.save()
                        user.set_password(password)
                        user.save()
                        return render(
                            request,
                            "coordinator/coordinator_reset_password.html",
                            {
                                "form": form,
                                "page_type": "reset_password",
                                "my_messages": {
                                    "success": "Password reset successfull. Download login credentions from below."
                                },
                            },
                        )
                    else:
                        return render(
                            request,
                            "coordinator/coordinator_reset_password.html",
                            {
                                "form": form,
                                "page_type": "reset_password",
                                "my_messages": {
                                    "error": "Parent does not belong to your organization."
                                },
                            },
                        )
                elif is_student(user):
                    user_student = StudentsInfo.objects.get(user=user)
                    if user_student.organization == organization:
                        password = random_password_generator()
                        user_student.first_password = password
                        user_student.password_changed = False
                        user_student.save()
                        user.set_password(password)
                        user.save()
                        return render(
                            request,
                            "coordinator/coordinator_reset_password.html",
                            {
                                "form": form,
                                "page_type": "reset_password",
                                "my_messages": {
                                    "success": "Password reset successfull. Download login credentions from below."
                                },
                            },
                        )
                    else:
                        return render(
                            request,
                            "coordinator/coordinator_reset_password.html",
                            {
                                "form": form,
                                "page_type": "reset_password",
                                "my_messages": {
                                    "error": "Student does not belong to your organization."
                                },
                            },
                        )
                else:
                    return render(
                        request,
                        "coordinator/coordinator_reset_password.html",
                        {
                            "form": form,
                            "page_type": "reset_password",
                            "my_messages": {
                                "error": "Provided user is not a teacher/parent/student."
                            },
                        },
                    )
            else:
                return render(
                    request,
                    "coordinator/coordinator_reset_password.html",
                    {
                        "form": form,
                        "page_type": "reset_password",
                        "my_messages": {"error": "Invalid username."},
                    },
                )
        else:
            return render(
                request,
                "coordinator/coordinator_reset_password.html",
                {"form": form, "page_type": "reset_password"},
            )


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_coordinator, login_url="accounts:forbidden")
@password_change_required
def coordinator_reset_password_teacher_download(request):
    output = io.BytesIO()
    wb = xlsxwriter.Workbook(output)
    bold = wb.add_format({"bold": True})
    credentials = wb.add_worksheet("Credentials")
    columns = [
        "Username",
        "Password",
    ]
    for col_num in range(len(columns)):
        credentials.write(0, col_num, columns[col_num], bold)
    teacher = TeacherInCharge.objects.filter(password_changed=False)
    for row_no, x in enumerate(teacher):
        credentials.write(row_no + 1, 0, x.user.username)
        credentials.write(row_no + 1, 1, x.first_password)
    wb.close()
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response[
        "Content-Disposition"
    ] = "attachment; filename=Teacher Login Credentials.xlsx"
    return response


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_coordinator, login_url="accounts:forbidden")
@password_change_required
def coordinator_reset_password_parent_download(request):
    output = io.BytesIO()
    wb = xlsxwriter.Workbook(output)
    bold = wb.add_format({"bold": True})
    credentials = wb.add_worksheet("Credentials")
    columns = [
        "Username",
        "Password",
    ]
    for col_num in range(len(columns)):
        credentials.write(0, col_num, columns[col_num], bold)
    parent = ParentsInfo.objects.filter(password_changed=False)
    for row_no, x in enumerate(parent):
        credentials.write(row_no + 1, 0, x.user.username)
        credentials.write(row_no + 1, 1, x.first_password)
    wb.close()
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response[
        "Content-Disposition"
    ] = "attachment; filename=Parent Login Credentials.xlsx"
    return response


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_coordinator, login_url="accounts:forbidden")
@password_change_required
def coordinator_reset_password_student_download(request):
    output = io.BytesIO()
    wb = xlsxwriter.Workbook(output)
    bold = wb.add_format({"bold": True})
    credentials = wb.add_worksheet("Credentials")
    columns = [
        "Username",
        "Password",
    ]
    for col_num in range(len(columns)):
        credentials.write(0, col_num, columns[col_num], bold)
    student = StudentsInfo.objects.filter(password_changed=False)
    for row_no, x in enumerate(student):
        credentials.write(row_no + 1, 0, x.user.username)
        credentials.write(row_no + 1, 1, x.first_password)
    wb.close()
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response[
        "Content-Disposition"
    ] = "attachment; filename=Student Login Credentials.xlsx"
    return response


@registration_data_cleanup
@redirect_to_dashboard
def forgot_password(request):
    if request.method == "GET":
        form = forgot_password_form()
        return render(request, "password/forgot_password.html", {"form": form})
    else:
        form = forgot_password_form(request.POST)
        context = {"form": form}
        if form.is_valid():
            username = form.cleaned_data["username"]
            group = form.cleaned_data["groups"]
            password = form.cleaned_data["password"]
            try:
                user = User.objects.get(username=username)
                if user.groups.filter(name=group).exists():
                    try:
                        validate_password(password)
                        if user.check_password(password):
                            form.add_error(
                                "password",
                                "Password entered is same as the previous one!",
                            )
                        else:
                            user.set_password(password)
                            user.save()
                            return redirect("accounts:password_changed")
                    except ValidationError as e:
                        form.add_error("password", e)
                else:
                    context["error"] = "Invalid Credentials."
            except User.DoesNotExist:
                context["error"] = "Invalid Credentials."
        return render(
            request,
            "password/forgot_password.html",
            context,
        )


@login_required(login_url="accounts:loginlink")
@user_passes_test(
    is_supercoordinator or is_coordinator or is_teacher or is_parent or is_student,
    login_url="accounts:forbidden",
)
def change_password(request):
    if request.method == "GET":
        form = change_password_form()
        return render(request, "password/change_password.html", {"form": form})
    else:
        form = change_password_form(request.POST)
        if form.is_valid():
            old_password = form.cleaned_data["old_password"]
            new_password = form.cleaned_data["password"]
            user = request.user
            if user.check_password(old_password):
                try:
                    validate_password(new_password)
                    if user.check_password(new_password):
                        form.add_error(
                            "password", "Password entered is same as the previous one!"
                        )
                    else:
                        user.set_password(new_password)
                        user.save()
                        if is_coordinator(user):
                            coord = CoordinatorInCharge.objects.get(user=user)
                            coord.password_changed = True
                            coord.first_password = ""
                            coord.save()
                        elif is_teacher(user):
                            teacher = TeacherInCharge.objects.get(user=user)
                            teacher.password_changed = True
                            teacher.first_password = ""
                            teacher.save()
                        elif is_parent(user):
                            parent = ParentsInfo.objects.get(user=user)
                            parent.password_changed = True
                            parent.first_password = ""
                            parent.save()
                        elif is_student(user):
                            student = StudentsInfo.objects.get(user=user)
                            student.password_changed = True
                            student.first_password = ""
                            student.save()
                        logout(request)
                        return redirect("accounts:password_changed")
                except ValidationError as e:
                    form.add_error("password", e)
            else:
                form.add_error("old_password", "Incorrect Password")
        return render(request, "password/change_password.html", {"form": form})


def password_changed(request):
    if request.method == "GET":
        return render(request, "password/password_changed.html", {})
    else:
        return redirect("accounts:loginlink")


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_parent or is_student, login_url="accounts:forbidden")
@password_change_required
@isActive("activity", "student")
def activity(request, user=None):
    if request.method == "GET":
        if user == None:
            user = request.user

        student = StudentsInfo.objects.get(user=user)
        startdate = FormDetails.objects.get(
            form=Form.objects.get(name="activity"), teacher=student.teacher, open=True
        ).start_timestamp
        if Activity.objects.filter(
            student=student, submission_timestamp__gte=startdate
        ).exists():
            draftForm = Activity.objects.get(
                student=StudentsInfo.objects.get(user=user),
                submission_timestamp__gte=startdate,
            )
            if draftForm.draft:
                mod = ActivityForm()
                temp = {}
                for name in Activity._meta.get_fields():
                    name = name.column
                    if name in mod.fields:
                        temp[name] = getattr(draftForm, name) or None
                form = ActivityForm(temp)
                formPre = getFormType("activity", student.teacher)
                return render(
                    request,
                    "activity/activity.html",
                    {"form": form, "formPre": formPre, "page_type": "student_activity"},
                )
            else:
                return redirect("../already_filled")
        # new form
        else:
            form = ActivityForm()
            formPre = getFormType("activity", student.teacher)
            return render(
                request,
                "activity/activity.html",
                {"form": form, "formPre": formPre, "page_type": "student_activity"},
            )
    # POST
    else:
        flag = False
        if user == None:
            flag = True
            user = request.user
        form = ActivityForm(request.POST)

        # valid form
        if form.is_valid():
            temp = createTempDict(request.POST)
            student = StudentsInfo.objects.get(user=user)
            startdate = FormDetails.objects.get(
                form=Form.objects.get(name="activity"),
                teacher=student.teacher,
                open=True,
            ).start_timestamp
            activityDraft(request)
            draftForm = Activity.objects.get(
                student=StudentsInfo.objects.get(user=user),
                submission_timestamp__gte=startdate,
            )
            if draftForm.draft:
                for name in Activity._meta.get_fields():
                    name = name.column
                    if name == "id" or name == "student_id" or name == "draft":
                        continue
                    elif name == "source_fruits_vegetables" or name == "grow_own_food":
                        my_list = "; ".join(ast.literal_eval(getattr(draftForm, name)))
                        setattr(draftForm, name, my_list)
                    elif name in temp:
                        setattr(draftForm, name, temp[name])

                draftForm.draft = False
                draftForm.submission_timestamp = datetime.now()
                draftForm.save()
            if flag:
                return redirect("accounts:student_dashboard")
            else:
                return redirect("accounts:parent_dashboard")
        # invalid form
        else:
            formPre = getFormType(
                "activity", StudentsInfo.objects.get(user=user).teacher
            )
            return render(
                request,
                "activity/activity.html",
                {"form": form, "formPre": formPre, "page_type": "student_activity"},
            )


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_parent or is_student, login_url="accounts:forbidden")
@password_change_required
def activityDraft(request):
    if "parent_dashboard" in request.META.get("HTTP_REFERER").split("/"):
        module = request.META.get("HTTP_REFERER").split("/")[-1]
        id = request.META.get("HTTP_REFERER").split("/")[-2]
        user = StudentsInfo.objects.get(pk=id).user
    else:
        module = request.META.get("HTTP_REFERER").split("/")[-2]
        user = request.user
    # for removing csrf field
    temp = createTempDict(request.POST)
    # checking if draft exists
    if not creatingOrUpdatingDraftsActivity(temp, user, "activity"):
        # creating new record
        form = Activity(**temp)
        form.student = StudentsInfo.objects.get(user=user)
        form.draft = True
        formType = getFormType("activity", StudentsInfo.objects.get(user=user).teacher)
        form.pre = 1 if formType == "PreTest" else 0
        form.submission_timestamp = datetime.now()
        if form.waist == "":
            form.waist = 0
        if form.weight == "":
            form.weight = 0
        if form.hip == "":
            form.hip = 0
        if form.height == "":
            form.height = 0
        form.save()
    return redirect(request.META.get("HTTP_REFERER"))


def creatingOrUpdatingDraftsActivity(temp, user, formName):
    student = StudentsInfo.objects.get(user=user)
    startdate = FormDetails.objects.get(
        form=Form.objects.get(name=formName), teacher=student.teacher, open=True
    ).start_timestamp
    if Activity.objects.filter(
        student=student, submission_timestamp__gte=startdate
    ).exists():
        draftForm = Activity.objects.get(
            student=StudentsInfo.objects.get(user=user),
            submission_timestamp__gte=startdate,
        )
        if draftForm.draft:
            # updating drafts
            for name in Activity._meta.get_fields():
                name = name.column
                if name == "id" or name == "student_id" or name == "draft":
                    continue
                if name in temp:
                    setattr(draftForm, name, temp[name])
                else:
                    setattr(draftForm, name, getattr(draftForm, name) or None)

            draftForm.submission_timestamp = datetime.now()
            if draftForm.waist == "":
                draftForm.waist = 0
            if draftForm.weight == "":
                draftForm.weight = 0
            if draftForm.hip == "":
                draftForm.hip = 0
            if draftForm.height == "":
                draftForm.height = 0
            draftForm.save()
            return True
    else:
        return False


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_parent, login_url="accounts:forbidden")
@password_change_required
@isActive("activity", "parent")
def parentActivity(request, id):
    user = StudentsInfo.objects.get(pk=id).user
    return activity(request, user)
