import io
import os
import openpyxl
import xlsxwriter
from datetime import datetime
from django.shortcuts import render, redirect
from django.http import HttpResponse, Http404
from django.contrib import messages
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required, user_passes_test
from .decorators import *
from .models import *
from .forms import *
from .helper_functions import *
from shared.encryption import EncryptionHelper
from django.utils import timezone
from django.conf import settings

encryptionHelper = EncryptionHelper()


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_teacher, login_url="accounts:forbidden")
@password_change_required
def teacherAllSessions(request):
    teacher = TeacherInCharge.objects.filter(user=request.user).first()
    objects = Teacher_Session.objects.filter(teacher=teacher)
    sessions = []
    for object in objects:
        sessions.append(object.session)
    upcoming_sessions = []
    open_sessions = []
    close_sessions = []
    open = False
    close = False
    upcoming = False
    for session in sessions:
        if session.start_date > timezone.now():
            upcoming_sessions.append(session)
            upcoming = True
        elif session.end_date == None:
            open_sessions.append(session)
            open = True
        else:
            close_sessions.append(session)
            close = True
    return render(
        request,
        "teacher/teacher_all_sessions.html",
        {
            "open_sessions": open_sessions,
            "upcoming_sessions": upcoming_sessions,
            "close_sessions": close_sessions,
            "open": open,
            "upcoming": upcoming,
            "close": close,
            "page_type": "teacher_all_sessions",
        },
    )


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_teacher, login_url="accounts:forbidden")
@password_change_required
def viewSessionStudents(request, id, open_id, my_messages=None):
    teacher = TeacherInCharge.objects.filter(user=request.user).first()
    session = Session.objects.filter(id=id).first()
    objects = Student_Session.objects.filter(session=session, teacher=teacher)
    students = []
    for object in objects:
        students.append(object.student)
    for student in students:
        student.fname = encryptionHelper.decrypt(student.fname)
        student.lname = encryptionHelper.decrypt(student.lname)
        student.unique_no = encryptionHelper.decrypt(student.unique_no)
    if my_messages != None:
        return render(
            request,
            "teacher/view_session_students.html",
            {
                "students": students,
                "session": session,
                "my_messages": my_messages,
                "page_type": "view_session_students",
                "open_id": open_id,
            },
        )
    else:
        return render(
            request,
            "teacher/view_session_students.html",
            {
                "students": students,
                "session": session,
                "page_type": "view_session_students",
                "open_id": open_id,
            },
        )


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_teacher, login_url="accounts:forbidden")
@password_change_required
def getSessionStudentsTemplate(request):
    output = io.BytesIO()
    wb = xlsxwriter.Workbook(output)
    ws = wb.add_worksheet("Session Students Data")
    columns = [
        "Student Username",
    ]
    for col_num in range(len(columns)):
        ws.write(0, col_num, columns[col_num])
    wb.close()
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = "attachment; filename=sessionStudentTemplate.xlsx"
    return response


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_teacher, login_url="accounts:forbidden")
@password_change_required
def addSessionStudents(request, id):
    if request.method == "GET":
        return render(
            request,
            "teacher/add_session_students.html",
            {"page_type": "add_session_students"},
        )
    else:
        try:
            excel_file = request.FILES["excel_file"]
            if excel_file.name[-4:] == "xlsx":
                try:
                    wb = openpyxl.load_workbook(excel_file)
                    studentSheet = wb["Session Students Data"]
                except:
                    return render(
                        request,
                        "teacher/add_session_students.html",
                        {
                            "page_type": "add_session_students",
                            "my_messages": {
                                "error": "Incorrect file uploaded, please use the template provided above."
                            },
                        },
                    )
            else:
                return render(
                    request,
                    "teacher/add_session_students.html",
                    {
                        "page_type": "add_session_students",
                        "my_messages": {
                            "error": "Incorrect file type, only .xlsx files are allowed!"
                        },
                    },
                )
        except:
            return render(
                request,
                "teacher/add_session_students.html",
                {
                    "page_type": "add_session_students",
                    "my_messages": {"error": "Sorry something went wrong!"},
                },
            )

        organization = (
            TeacherInCharge.objects.filter(user=request.user).first().organization
        )
        breaking = error = False
        error_message = ""
        student_data = []
        for row_no, row in enumerate(studentSheet.iter_rows()):
            if breaking == True:
                break
            if row_no == 0:
                continue
            for cell in row:
                if cell.column_letter == "A":
                    if cell.value == None:
                        breaking = True
                        break
                    else:
                        if User.objects.filter(username=cell.value).exists():
                            user = User.objects.filter(username=cell.value).first()
                            if is_student(user):
                                student_user = StudentsInfo.objects.filter(
                                    user=user
                                ).first()
                                if student_user.organization == organization:
                                    student_data.append(student_user)
                                else:
                                    breaking = error = True
                                    error_message = (
                                        "Student at row number "
                                        + str(row_no + 1)
                                        + " does not belong to your organization"
                                    )
                                    break
                            else:
                                breaking = error = True
                                error_message = (
                                    "User at row number "
                                    + str(row_no + 1)
                                    + " is not a student"
                                )
                                break
                        else:
                            breaking = error = True
                            error_message = "Invalid username at row number " + str(
                                row_no + 1
                            )
                            break

        if error == True:
            return render(
                request,
                "teacher/add_session_students.html",
                {
                    "page_type": "add_session_students",
                    "my_messages": {"error": error_message},
                },
            )

        session = Session.objects.filter(id=id).first()
        student_added = 0
        student_exists = 0
        for student_user in student_data:
            if student_user.session != None:
                student_exists += 1
            else:
                teacher = TeacherInCharge.objects.filter(user=request.user).first()
                student_user.session = session
                student_user.teacher = TeacherInCharge.objects.filter(
                    user=request.user
                ).first()
                student_session = Student_Session()
                student_session.session = session
                student_session.student = student_user
                student_session.teacher = teacher
                student_user.save()
                student_session.save()
                student_added += 1

        if student_added == 0:
            my_messages = {
                "success": "Registration successful. Already registered: "
                + str(student_exists)
            }
        elif student_exists == 0:
            my_messages = {
                "success": "Registration successful. Newly registered: "
                + str(student_added)
            }
        else:
            my_messages = {
                "success": "Registration successful. Already registered: "
                + str(student_exists)
                + ", Newly registered: "
                + str(student_added)
            }

        return viewSessionStudents(request, id, 1, my_messages)


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_teacher, login_url="accounts:forbidden")
@password_change_required
def addSessionStudentsList(request, id):
    if request.method == "GET":
        students = StudentsInfo.objects.filter(session=None, teacher=None)
        for student in students:
            student.fname = encryptionHelper.decrypt(student.fname)
            student.lname = encryptionHelper.decrypt(student.lname)
            student.unique_no = encryptionHelper.decrypt(student.unique_no)
        return render(
            request,
            "teacher/add_session_students_list.html",
            {
                "page_type": "add_session_students_list",
                "students": students,
            },
        )
    else:
        students_id_list = request.POST.getlist("chk")
        session = Session.objects.filter(id=id).first()
        teacher = TeacherInCharge.objects.filter(user=request.user).first()
        for s_id in students_id_list:
            student = StudentsInfo.objects.filter(id=s_id).first()
            student.session = session
            student.teacher = teacher
            student_session = Student_Session()
            student_session.session = session
            student_session.student = student
            student_session.teacher = teacher
            student_session.save()
            student.save()

        return redirect("accounts:view_session_students", id, 1)


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_teacher, login_url="accounts:forbidden")
@password_change_required
def removeSessionStudent(request, session_id, student_id):
    student = StudentsInfo.objects.filter(id=student_id).first()
    student.session = None
    student.teacher = None
    Student_Session.objects.filter(student=student).delete()
    student.save()
    my_messages = {"success": "Student removed from session successfully"}
    return viewSessionStudents(request, session_id, 1, my_messages)


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_teacher, login_url="accounts:forbidden")
@password_change_required
def viewSessionForms(request, id):
    teacher = TeacherInCharge.objects.get(user=request.user)
    teachersession = Session.objects.filter(id=id).first()
    total_students = StudentsInfo.objects.filter(
        session=teachersession, teacher=teacher
    )
    results = []
    closed_sessions = FormDetails.objects.filter(
        teacher=teacher, session=teachersession, open=False
    )

    for session in closed_sessions:
        temp_list = [session.form, session.start_timestamp, session.end_timestamp]
        if session.pre:
            temp_list.append("Pre Test")
        else:
            temp_list.append("Post Test")
        count = 0
        for student in total_students:
            if ModuleOne.objects.filter(
                student=student,
                submission_timestamp__gte=session.start_timestamp,
                submission_timestamp__lte=session.end_timestamp,
            ).exists():
                draftForm = ModuleOne.objects.filter(
                    student=student,
                    submission_timestamp__gte=session.start_timestamp,
                    submission_timestamp__lte=session.end_timestamp,
                ).first()
                if not draftForm.draft:
                    count += 1
            elif Physique.objects.filter(
                student=student,
                submission_timestamp__gte=session.start_timestamp,
                submission_timestamp__lte=session.end_timestamp,
            ).exists():
                draftForm = Physique.objects.filter(
                    student=student,
                    submission_timestamp__gte=session.start_timestamp,
                    submission_timestamp__lte=session.end_timestamp,
                ).first()
                if not draftForm.draft:
                    count += 1

        temp_list.append(count)
        temp_list.append(len(total_students))
        temp_list.append(session.id)
        results.append(temp_list)
    open_sessions = FormDetails.objects.filter(
        teacher=teacher, session=teachersession, open=True
    )
    results2 = []
    for session in open_sessions:
        temp_list = [session.form, session.start_timestamp]
        if session.pre:
            temp_list.append("Pre Test")
        else:
            temp_list.append("Post Test")
        count = 0
        for student in total_students:
            if ModuleOne.objects.filter(
                student=student, submission_timestamp__gte=session.start_timestamp
            ).exists():
                draftForm = ModuleOne.objects.filter(
                    student=student, submission_timestamp__gte=session.start_timestamp
                ).first()
                if not draftForm.draft:
                    count += 1
            elif Physique.objects.filter(
                student=student, submission_timestamp__gte=session.start_timestamp
            ).exists():
                draftForm = Physique.objects.filter(
                    student=student, submission_timestamp__gte=session.start_timestamp
                ).first()
                if not draftForm.draft:
                    count += 1

        temp_list.append(count)
        temp_list.append(len(total_students))
        temp_list.append(session.id)
        results2.append(temp_list)

    return render(
        request,
        "teacher/view_session_forms.html",
        {
            "results": results,
            "results2": results2,
            "page_type": "view_session_forms",
            "session": teachersession,
        },
    )


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_teacher, login_url="accounts:forbidden")
@password_change_required
def getFormDetails(request, id):
    form = FormDetails.objects.get(pk=id)
    teacher = form.teacher
    session_students = Student_Session.objects.filter(teacher = teacher, session = form.session)
    total_students = []
    for session_student in session_students:
        total_students.append(session_student.student)

    filled_students = []
    not_filled_students = []
    if not form.open:
        form_open = False
        temp_list = [form.form, form.start_timestamp, form.end_timestamp]
    else:
        form_open = True
        temp_list = [form.form, form.start_timestamp]

    if form.pre:
        temp_list.append("Pre Test")
    else:
        temp_list.append("Post Test")

    if form.form.name == "moduleOne":
        for student in total_students:
            temp = []
            if form.open:
                if ModuleOne.objects.filter(
                    student=student, submission_timestamp__gte=form.start_timestamp
                ).exists():
                    draftForm = ModuleOne.objects.filter(
                        student=student, submission_timestamp__gte=form.start_timestamp
                    ).first()

                    if draftForm.draft:
                        temp.append(
                            encryptionHelper.decrypt(student.fname)
                            + " "
                            + encryptionHelper.decrypt(student.lname)
                        )
                        temp.append("-")
                        not_filled_students.append(temp)
                    else:
                        temp.append(
                            encryptionHelper.decrypt(student.fname)
                            + " "
                            + encryptionHelper.decrypt(student.lname)
                        )
                        temp.append(draftForm.submission_timestamp)
                        filled_students.append(temp)
                else:
                    temp.append(
                        encryptionHelper.decrypt(student.fname)
                        + " "
                        + encryptionHelper.decrypt(student.lname)
                    )
                    temp.append("-")
                    not_filled_students.append(temp)

            else:
                if ModuleOne.objects.filter(
                    student=student,
                    submission_timestamp__gte=form.start_timestamp,
                    submission_timestamp__lte=form.end_timestamp,
                ).exists():
                    submitted_form = ModuleOne.objects.filter(
                        student=student, submission_timestamp__gte=form.start_timestamp
                    ).first()
                    temp.append(
                        encryptionHelper.decrypt(student.fname)
                        + " "
                        + encryptionHelper.decrypt(student.lname)
                    )
                    temp.append(submitted_form.submission_timestamp)
                    filled_students.append(temp)
                else:
                    temp.append(
                        encryptionHelper.decrypt(student.fname)
                        + " "
                        + encryptionHelper.decrypt(student.lname)
                    )
                    temp.append("-")
                    not_filled_students.append(temp)

    elif form.form.name == "physique":
        for student in total_students:
            temp = []
            if Physique.objects.filter(
                    student=student, submission_timestamp__gte=form.start_timestamp
                ).exists():
                    draftForm = Physique.objects.filter(
                        student=student, submission_timestamp__gte=form.start_timestamp
                    ).first()

                    if draftForm.draft:
                        temp.append(
                            encryptionHelper.decrypt(student.fname)
                            + " "
                            + encryptionHelper.decrypt(student.lname)
                        )
                        temp.append("-")
                        not_filled_students.append(temp)
                    else:
                        temp.append(
                            encryptionHelper.decrypt(student.fname)
                            + " "
                            + encryptionHelper.decrypt(student.lname)
                        )
                        temp.append(draftForm.submission_timestamp)
                        filled_students.append(temp)
            else:
                temp.append(encryptionHelper.decrypt(student.fname) + " " + encryptionHelper.decrypt(student.lname))
                temp.append("-")
                not_filled_students.append(temp)


    return render(
        request,
        "teacher/teacher_dashboard_getDetails.html",
        {
            "result": temp_list,
            "filled_students": filled_students,
            "not_filled_students": not_filled_students,
            "open": form_open,
        },
    )


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_teacher, login_url="accounts:forbidden")
@password_change_required
def manageForms(request, id):
    if request.method == "GET":
        physique = {}
        moduleOne = {}
        moduleTwo = {}
        moduleThree = {}
        teacher = TeacherInCharge.objects.get(user=request.user)
        session = Session.objects.filter(id=id).first()
        form = Form.objects.get(name="moduleOne")
        if FormDetails.objects.filter(
            form=form, teacher=teacher, session=session
        ).exists():
            form = (
                FormDetails.objects.filter(
                    form=form, teacher=teacher, session=session, open=True
                )
                .order_by("-start_timestamp")
                .first()
            )
            if form:
                if form.pre:
                    moduleOne["pre"] = True
                    moduleOne["post"] = False

                else:
                    moduleOne["post"] = True
                    moduleOne["pre"] = False

        physique_form = Form.objects.get(name="physique")
        if FormDetails.objects.filter(
            form=physique_form, teacher=teacher, session=session
        ).exists():
            form2 = (
                FormDetails.objects.filter(
                    form=physique_form, teacher=teacher, session=session, open=True
                )
                .order_by("-start_timestamp")
                .first()
            )

        return render(
            request,
            "teacher/manage_forms_teacher.html",
            {
                "physique": physique,
                "moduleOne": moduleOne,
                "moduleTwo": moduleTwo,
                "moduleThree": moduleThree,
                "page_type": "manage_forms",
            },
        )
    else:

        if "moduleOne" in request.POST:
            module_one_pre = request.POST.get("module_one_pre", False)
            module_one_post = request.POST.get("module_one_post", False)
            if module_one_pre == "on" and module_one_post == "on":
                messages.error(request, "Cannot select both PreTest and PostTest")
                return redirect("accounts:manage_forms")

            form = Form.objects.get(name="moduleOne")
            teacher = TeacherInCharge.objects.get(user=request.user)
            session = Session.objects.filter(id=id).first()
            if module_one_pre == "on":
                if not FormDetails.objects.filter(
                    form=form, teacher=teacher, session=session, pre=True, open=True
                ).exists():
                    update = FormDetails(
                        form=form,
                        teacher=teacher,
                        session=session,
                        pre=True,
                        open=True,
                        start_timestamp=datetime.now(),
                    )
                    update.save()
            else:
                if FormDetails.objects.filter(
                    form=form, teacher=teacher, session=session, pre=True, open=True
                ).exists():
                    update = FormDetails.objects.filter(
                        form=form, teacher=teacher, session=session, pre=True, open=True
                    ).first()
                    update.open = False
                    update.end_timestamp = datetime.now()
                    session = Session.objects.filter(id=id).first()
                    teacher = TeacherInCharge.objects.filter(user=request.user).first()
                    total_students = StudentsInfo.objects.filter(
                        session=session, teacher=teacher
                    )
                    for student in total_students:
                        if ModuleOne.objects.filter(
                            student=student,
                            submission_timestamp__gte=update.start_timestamp,
                            submission_timestamp__lte=update.end_timestamp,
                            draft=True,
                            pre=True,
                        ).exists():
                            draftForm = ModuleOne.objects.filter(
                                student=student,
                                submission_timestamp__gte=update.start_timestamp,
                                submission_timestamp__lte=update.end_timestamp,
                                draft=True,
                                pre=True,
                            ).first()
                            draftForm.delete()
                    update.save()

            if module_one_post == "on":
                if not FormDetails.objects.filter(
                    form=form, teacher=teacher, session=session, pre=False, open=True
                ).exists():
                    update = FormDetails(
                        form=form,
                        teacher=teacher,
                        session=session,
                        pre=False,
                        open=True,
                        start_timestamp=datetime.now(),
                    )
                    update.save()
            else:
                if FormDetails.objects.filter(
                    form=form, teacher=teacher, session=session, pre=False, open=True
                ).exists():
                    update = FormDetails.objects.filter(
                        form=form,
                        teacher=teacher,
                        session=session,
                        pre=False,
                        open=True,
                    ).first()
                    update.open = False
                    update.end_timestamp = datetime.now()
                    teacher = TeacherInCharge.objects.get(user=request.user)
                    session = Session.objects.filter(id=id).first()
                    total_students = StudentsInfo.objects.filter(
                        session=session, teacher=teacher
                    )
                    for student in total_students:
                        if ModuleOne.objects.filter(
                            student=student,
                            submission_timestamp__gte=update.start_timestamp,
                            submission_timestamp__lte=update.end_timestamp,
                            draft=True,
                            pre=False,
                        ).exists():
                            draftForm = ModuleOne.objects.filter(
                                student=student,
                                submission_timestamp__gte=update.start_timestamp,
                                submission_timestamp__lte=update.end_timestamp,
                                draft=True,
                                pre=False,
                            ).first()
                            draftForm.delete()
                    update.save()

        elif "physique" in request.POST:
            physique_pre = request.POST.get("physique_pre", False) 
            form = Form.objects.get(name="physique")
            teacher = TeacherInCharge.objects.get(user=request.user)
            session = Session.objects.filter(id=id).first()
            total_students = StudentsInfo.objects.filter(
                session=session, teacher=teacher
            )

            # if physique_pre == "on":
            if not FormDetails.objects.filter(
                    form=form, teacher=teacher, session=session, pre=True, open=True
                ).exists():
                    update = FormDetails(
                        form=form,
                        teacher=teacher,
                        session=session,
                        pre=True,
                        open=True,
                        start_timestamp=datetime.now(),
                    )
                    update.save()
            else:
                if FormDetails.objects.filter(
                    form=form, teacher=teacher, session=session, pre=True, open=True
                ).exists():
                    update = FormDetails.objects.filter(
                        form=form, teacher=teacher, session=session, pre=True, open=True
                    ).first()
                    update.open = False
                    update.end_timestamp = datetime.now()
                    teacher = TeacherInCharge.objects.get(user=request.user)
                    session = Session.objects.filter(id=id).first()
                    total_students = StudentsInfo.objects.filter(
                        session=session, teacher=teacher
                    )
                    for student in total_students:
                        if Physique.objects.filter(
                            student=student,
                            submission_timestamp__gte=update.start_timestamp,
                            submission_timestamp__lte=update.end_timestamp,
                            draft=True,
                            pre=True,
                        ).exists():
                            draftForm = Physique.objects.filter(
                                student=student,
                                submission_timestamp__gte=update.start_timestamp,
                                submission_timestamp__lte=update.end_timestamp,
                                draft=True,
                                pre=True,
                            ).first()
                            draftForm.delete()
                    update.save()

            # if physique_post == "on":
            #     if not FormDetails.objects.filter(
            #         form=form, teacher=teacher, session=session, pre=False, open=True
            #     ).exists():
            #         update = FormDetails(
            #             form=form,
            #             teacher=teacher,
            #             session=session,
            #             pre=False,
            #             open=True,
            #             start_timestamp=datetime.now(),
            #         )
            #         update.save()
            # else:
            #     if FormDetails.objects.filter(
            #         form=form, teacher=teacher, session=session, pre=False, open=True
            #     ).exists():
            #         update = FormDetails.objects.filter(
            #             form=form,
            #             teacher=teacher,
            #             session=session,
            #             pre=False,
            #             open=True,
            #         ).first()
            #         update.open = False
            #         update.end_timestamp = datetime.now()
            #         teacher = TeacherInCharge.objects.get(user=request.user)
            #         session = Session.objects.filter(id=id).first()
            #         total_students = StudentsInfo.objects.filter(
            #             session=session, teacher=teacher
            #         )
            #         for student in total_students:
            #             if Physique.objects.filter(
            #                 student=student,
            #                 submission_timestamp__gte=update.start_timestamp,
            #                 submission_timestamp__lte=update.end_timestamp,
            #                 draft=True,
            #                 pre=False,
            #             ).exists():
            #                 draftForm = Physique.objects.filter(
            #                     student=student,
            #                     submission_timestamp__gte=update.start_timestamp,
            #                     submission_timestamp__lte=update.end_timestamp,
            #                     draft=True,
            #                     pre=False,
            #                 ).first()
            #                 draftForm.delete()
            #         update.save()

        module_two_pre = request.POST.get("module_two_pre", False)
        module_two_post = request.POST.get("module_two_post", False)
        module_three_pre = request.POST.get("module_three_pre", False)
        module_three_post = request.POST.get("module_three_post", False)

        return redirect("accounts:manage_forms", id)


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_teacher, login_url="accounts:forbidden")
@password_change_required
def getTemplate(request):
    file_path = os.path.join(
        settings.BASE_DIR, "accounts/Bulk Registration Template.xlsx"
    )
    if os.path.exists(file_path):
        with open(file_path, "rb") as fh:
            response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
            response["Content-Disposition"] = "inline; filename=" + os.path.basename(
                file_path
            )
            return response
    raise Http404


# Work in progress
@login_required(login_url="accounts:loginlink")
@user_passes_test(is_teacher, login_url="accounts:forbidden")
@password_change_required
def bulkRegister(request):
    if request.method == "GET":
        return render(
            request,
            "teacher/bulkregistration.html",
            {"page_type": "bulk_register"},
        )
    else:
        try:
            excel_file = request.FILES["excel_file"]
            if excel_file.name[-4:] == "xlsx":
                try:
                    wb = openpyxl.load_workbook(excel_file)
                    parentSheet = wb["Parents Data"]
                    studentSheet = wb["Students Data"]
                except:
                    return render(
                        request,
                        "teacher/bulkregistration.html",
                        {
                            "page_type": "bulk_register",
                            "my_messages": {
                                "error": "Incorrect file uploaded, please use the template provided above."
                            },
                        },
                    )
            else:
                return render(
                    request,
                    "teacher/bulkregistration.html",
                    {
                        "page_type": "bulk_register",
                        "my_messages": {
                            "error": "Incorrect file type, only .xlsx files are allowed!"
                        },
                    },
                )
        except:
            return render(
                request,
                "teacher/bulkregistration.html",
                {
                    "page_type": "bulk_register",
                    "my_messages": {"error": "Sorry something went wrong!"},
                },
            )

        breaking = error = False
        error_message = ""
        parent_data = []
        for row_no, row in enumerate(parentSheet.iter_rows()):
            if breaking == True:
                break
            if row_no == 0:
                continue
            row_data = []
            for cell in row:
                if cell.column_letter == "A":
                    if cell.value == None:
                        breaking = True
                        break
                    elif cell.value != row_no:
                        breaking = error = True
                        error_message = (
                            "There something wrong in the parentId column in parent data sheet at row number "
                            + str(row_no + 1)
                        )
                        break
                elif cell.column_letter == "B":
                    if cell.value == "-" or cell.value == None:
                        row_data.append(None)
                    else:
                        if User.objects.filter(username=str(cell.value)).exists():
                            row_data.append(cell.value)
                            break
                        else:
                            breaking = error = True
                            error_message = (
                                "Parent user with the following username at row number "
                                + str(row_no + 1)
                                + " does not exist"
                            )
                            break
                elif cell.column_letter == "C":
                    if cell.value == None:
                        breaking = True
                        break
                    if not valid_name(cell.value):
                        breaking = error = True
                        error_message = "Invalid parent name at row number " + str(
                            row_no + 1
                        )
                        break
                    else:
                        row_data.append(cell.value)
                elif cell.column_letter == "D":
                    if cell.value != None and not valid_email(cell.value):
                        breaking = error = True
                        error_message = "Invalid parent email at row number " + str(
                            row_no + 1
                        )
                        break
                    else:
                        row_data.append(cell.value)
                elif cell.column_letter == "E":
                    if cell.value != None and not valid_mobile_no(cell.value):
                        breaking = error = True
                        error_message = (
                            "Invalid parent mobile number at row number "
                            + str(row_no + 1)
                        )
                        break
                    else:
                        row_data.append(cell.value)
                elif cell.column_letter == "F":
                    if cell.value == None:
                        breaking = error = True
                        error_message = "Parent's gender missing at row number " + str(
                            row_no + 1
                        )
                        break
                    else:
                        if check_gender(str(cell.value)):
                            row_data.append(cell.value)
                        else:
                            breaking = error = True
                            error_message = (
                                "Invalid gender input for parents data at row number  "
                                + str(row_no + 1)
                                + ", please read the instructions"
                            )
                            break
                elif cell.column_letter == "G":
                    if cell.value == None:
                        breaking = error = True
                        error_message = (
                            "Parent's date of birth missing at row number "
                            + str(row_no + 1)
                        )
                        break
                    else:
                        if re.match(
                            "^[0-9]{2}/[0-9]{2}/[0-9]{4}$", str(cell.value)
                        ) and valid_date(str(cell.value)):
                            if is_adult_func(str(cell.value)) == "True":
                                row_data.append(cell.value)
                            else:
                                breaking = error = True
                                error_message = (
                                    "Parent at row number "
                                    + str(row_no + 1)
                                    + " is not an adult"
                                )
                            break
                        else:
                            breaking = error = True
                            error_message = (
                                "Invalid parent's date of birth at row number "
                                + str(row_no + 1)
                            )
                            break
                elif cell.column_letter == "H":
                    if cell.value == None:
                        breaking = error = True
                        error_message = "Parent's pincode missing at row number " + str(
                            row_no + 1
                        )
                        break
                    else:
                        if not valid_pincode(cell.value):
                            breaking = error = True
                            error_message = (
                                "Invalid parent's pincode at row number "
                                + str(row_no + 1)
                            )
                            break
                        else:
                            row_data.append(cell.value)
                elif cell.column_letter == "I":
                    if cell.value == None:
                        breaking = error = True
                        error_message = "Parent's state missing at row number " + str(
                            row_no + 1
                        )
                        break
                    else:
                        temp = check_state_city(True, 0, str(cell.value))
                        if temp[0]:
                            row_data.append(cell.value)
                            row_data.append(temp[1])
                        else:
                            breaking = error = True
                            error_message = (
                                "Invalid parent's state at row number "
                                + str(row_no + 1)
                            )
                            break
                elif cell.column_letter == "J":
                    if cell.value == None:
                        breaking = error = True
                        error_message = "Parent's city missing at row number " + str(
                            row_no + 1
                        )
                        break
                    else:
                        if check_state_city(False, row_data[-1], str(cell.value)):
                            row_data.pop()
                            row_data.append(cell.value)
                        else:
                            breaking = error = True
                            error_message = (
                                "Invalid parent's city at row number " + str(row_no + 1)
                            )
                            break
                elif cell.column_letter == "K":
                    if cell.value == None:
                        breaking = error = True
                        error_message = "Parent's address missing at row number " + str(
                            row_no + 1
                        )
                        break
                    else:
                        row_data.append(cell.value)
                elif cell.column_letter == "L":
                    if cell.value == None:
                        breaking = error = True
                        error_message = (
                            "Parent's religion missing at row number " + str(row_no + 1)
                        )
                        break
                    else:
                        if check_religion(str(cell.value)):
                            row_data.append(cell.value)
                        else:
                            breaking = error = True
                            error_message = (
                                "Invalid parent's religion at row number "
                                + str(row_no + 1)
                            )
                            break
                elif cell.column_letter == "M":
                    if cell.value == None:
                        breaking = error = True
                        error_message = (
                            "Parent's type of family missing at row number "
                            + str(row_no + 1)
                        )
                        break
                    else:
                        if check_type_of_fam(str(cell.value)):
                            row_data.append(cell.value)
                        else:
                            breaking = error = True
                            error_message = (
                                "Invalid parent's type of family at row number "
                                + str(row_no + 1)
                            )
                            break
                elif cell.column_letter == "N":
                    if cell.value == None:
                        breaking = error = True
                        error_message = (
                            "Parent's no of family members missing at row number "
                            + str(row_no + 1)
                        )
                        break
                    else:
                        try:
                            int(cell.value)
                        except:
                            breaking = error = True
                            error_message = (
                                "Invalid parent's no of family members at row number "
                                + str(row_no + 1)
                            )
                            break
                        row_data.append(cell.value)
                elif cell.column_letter == "O":
                    if cell.value == None:
                        breaking = error = True
                        error_message = (
                            "Parent's children count missing at row number "
                            + str(row_no + 1)
                        )
                        break
                    else:
                        try:
                            int(cell.value)
                        except:
                            breaking = error = True
                            error_message = (
                                "Invalid parent's children count at row number "
                                + str(row_no + 1)
                            )
                            break
                        row_data.append(cell.value)
                elif cell.column_letter == "P":
                    if cell.value == None:
                        breaking = error = True
                        error_message = (
                            "Parent's education missing at row number "
                            + str(row_no + 1)
                        )
                        break
                    else:
                        if check_education(str(cell.value)):
                            row_data.append(cell.value)
                        else:
                            breaking = error = True
                            error_message = (
                                "Invalid parent's education at row number "
                                + str(row_no + 1)
                            )
                            break
                elif cell.column_letter == "Q":
                    if cell.value == None:
                        breaking = error = True
                        error_message = (
                            "Parent's occupation missing at row number "
                            + str(row_no + 1)
                        )
                        break
                    else:
                        if check_occupation(str(cell.value)):
                            row_data.append(cell.value)
                        else:
                            breaking = error = True
                            error_message = (
                                "Invalid parent's occupation at row number "
                                + str(row_no + 1)
                            )
                            break
            parent_data.append(row_data)

        if error == True:
            return render(
                request,
                "teacher/bulkregistration.html",
                {
                    "page_type": "bulk_register",
                    "my_messages": {"error": error_message},
                },
            )

        breaking = error = False
        error_message = ""
        student_data = []
        for row_no, row in enumerate(studentSheet.iter_rows()):
            if breaking == True:
                break
            if row_no == 0:
                continue
            row_data = []
            for cell in row:
                if cell.column_letter == "A":
                    if cell.value == None:
                        breaking = True
                        break
                    if not valid_name(cell.value):
                        breaking = error = True
                        error_message = "Invalid student name at row number " + str(
                            row_no + 1
                        )
                        break
                    else:
                        row_data.append(cell.value)
                elif cell.column_letter == "B":
                    if cell.value != None and not valid_email(cell.value):
                        breaking = error = True
                        error_message = "Invalid student email at row number " + str(
                            row_no + 1
                        )
                        break
                    row_data.append(cell.value)
                elif cell.column_letter == "C":
                    if cell.value != None and not valid_mobile_no(cell.value):
                        breaking = error = True
                        error_message = (
                            "Invalid student mobile number at row number "
                            + str(row_no + 1)
                        )
                        break
                    row_data.append(cell.value)
                elif cell.column_letter == "D":
                    if cell.value == None:
                        breaking = error = True
                        error_message = "Student's gender missing at row number " + str(
                            row_no + 1
                        )
                        break
                    else:
                        if check_gender(str(cell.value)):
                            row_data.append(cell.value)
                        else:
                            breaking = error = True
                            error_message = (
                                "Invalid gender input for students data at row number  "
                                + str(row_no + 1)
                                + ", please read the instructions"
                            )
                            break
                elif cell.column_letter == "E":
                    if cell.value == None:
                        breaking = error = True
                        error_message = (
                            "Student's date of birth missing at row number "
                            + str(row_no + 1)
                        )
                        break
                    else:
                        if re.match(
                            "^[0-9]{2}/[0-9]{2}/[0-9]{4}$", str(cell.value)
                        ) and valid_date(str(cell.value)):
                            dob = cell.value[6:] + cell.value[:2] + cell.value[3:5]
                            if valid_dob(dob):
                                row_data.append(cell.value)
                            else:
                                breaking = error = True
                                error_message = (
                                    "Student at row number "
                                    + str(row_no + 1)
                                    + " is not above the age of 5"
                                )
                            break
                        else:
                            breaking = error = True
                            error_message = (
                                "Invalid student's date of birth at row number "
                                + str(row_no + 1)
                            )
                            break
                elif cell.column_letter == "F":
                    if cell.value == None:
                        breaking = error = True
                        error_message = (
                            "Student's pincode missing at row number " + str(row_no + 1)
                        )
                        break
                    else:
                        if not valid_pincode(cell.value):
                            breaking = error = True
                            error_message = (
                                "Invalid student's pincode at row number "
                                + str(row_no + 1)
                            )
                            break
                        else:
                            row_data.append(cell.value)
                elif cell.column_letter == "G":
                    if cell.value == None:
                        breaking = error = True
                        error_message = "Student's state missing at row number " + str(
                            row_no + 1
                        )
                        break
                    else:
                        temp = check_state_city(True, 0, str(cell.value))
                        if temp[0]:
                            row_data.append(cell.value)
                            row_data.append(temp[1])
                        else:
                            breaking = error = True
                            error_message = (
                                "Invalid student's state at row number "
                                + str(row_no + 1)
                            )
                            break
                elif cell.column_letter == "H":
                    if cell.value == None:
                        breaking = error = True
                        error_message = "Student's city missing at row number " + str(
                            row_no + 1
                        )
                        break
                    else:
                        if check_state_city(False, row_data[-1], str(cell.value)):
                            row_data.pop()
                            row_data.append(cell.value)
                        else:
                            breaking = error = True
                            error_message = (
                                "Invalid student's city at row number "
                                + str(row_no + 1)
                            )
                            break
                elif cell.column_letter == "I":
                    if cell.value == None:
                        breaking = error = True
                        error_message = (
                            "Student's address missing at row number " + str(row_no + 1)
                        )
                        break
                    else:
                        row_data.append(cell.value)
                elif cell.column_letter == "J":
                    if cell.value == None:
                        breaking = error = True
                        error_message = (
                            "Student's registration no missing at row number "
                            + str(row_no + 1)
                        )
                        break
                    else:
                        row_data.append(cell.value)
                elif cell.column_letter == "K":
                    if cell.value == None:
                        dob = row_data[-6]
                        if is_adult_func(dob) == "True":
                            row_data.append("ADULT")
                        else:
                            breaking = error = True
                            error_message = (
                                "Student at row number "
                                + str(row_no + 1)
                                + " is not an adult, please provide parentId for the student"
                            )
                            break
                    else:
                        try:
                            parentId = int(cell.value)
                            if parentId <= len(parent_data):
                                row_data.append(cell.value)
                            else:
                                breaking = error = True
                                error_message = (
                                    "Invalid parentId for student data at row number "
                                    + str(row_no + 1)
                                )
                                break
                        except:
                            breaking = error = True
                            error_message = (
                                "Invalid parentId for student data at row number "
                                + str(row_no + 1)
                            )
                            break
            student_data.append(row_data)

        if error == True:
            return render(
                request,
                "teacher/bulkregistration.html",
                {
                    "page_type": "bulk_register",
                    "my_messages": {"error": error_message},
                },
            )

        # for index, row in enumerate(parent_data):
        #     if index == 0:
        #         continue
        #     # creating parent user
        #     skipparent = False
        #     parentData = ParentsInfo.objects.all()
        #     for parent in parentData:
        #         if encryptionHelper.decrypt(parent.email) == encryptionHelper.decrypt(
        #             row[0]
        #         ):
        #             skipparent = True
        #             break

        #     if not skipparent:
        #         password = "".join(
        #             random.choices(string.ascii_lowercase + string.digits, k=8)
        #         )
        #         parentUser = User(username=row[2])
        #         parentUser.set_password(password)
        #         parentUser.save()
        #         parent_group = Group.objects.get(name="Parents")
        #         parentUser.groups.add(parent_group)
        #         parentUser.save()

        #         # getting data from db for foreign keys
        #         city = City.objects.filter(city__icontains=row[9]).first()
        #         state = State.objects.filter(state__icontains=row[10]).first()
        #         education = Education.objects.filter(
        #             education__icontains=row[11]
        #         ).first()
        #         occupation = Occupation.objects.filter(
        #             occupation__icontains=row[12]
        #         ).first()
        #         religion = ReligiousBelief.objects.filter(
        #             religion__icontains=row[13]
        #         ).first()
        #         familyType = FamilyType.objects.filter(
        #             family__icontains=row[14]
        #         ).first()
        #         # creating parent
        #         parent = ParentsInfo(
        #             email=row[0],
        #             name=row[1],
        #             gender=row[3],
        #             age=row[4],
        #             address=row[5],
        #             pincode=row[6],
        #             no_of_family_members=row[7],
        #             children_count=row[8],
        #             city=city,
        #             state=state,
        #             edu=education,
        #             occupation=occupation,
        #             religion=religion,
        #             type_of_family=familyType,
        #             first_password=password,
        #         )
        #         parent.user = parentUser
        #         parent.save()

        # for index, row in enumerate(student_data):
        #     if index == 0:
        #         continue
        #     # creating student user
        #     skipstudent = StudentsInfo.objects.filter(rollno=row[3]).first()
        #     if not skipstudent:
        #         password = "".join(
        #             random.choices(string.ascii_lowercase + string.digits, k=8)
        #         )
        #         studentUser = User(username=row[1])
        #         studentUser.set_password(password)
        #         studentUser.save()
        #         student_group = Group.objects.get(name="Students")
        #         studentUser.groups.add(student_group)
        #         studentUser.save()

        #         # getting data from db for foreign keys
        #         parentData = ParentsInfo.objects.all()
        #         for tempparent in parentData:
        #             if encryptionHelper.decrypt(tempparent.email) == row[7]:
        #                 parent = tempparent

        #         organization = Organization.objects.filter(
        #             name__icontains=row[6]
        #         ).first()
        #         teacher = TeacherInCharge.objects.filter(user=request.user).first()
        #         # creating student
        #         dob = datetime.strptime(row[5], "%Y-%m-%d %H:%M:%S").strftime(
        #             "%Y-%m-%d"
        #         )
        #         student = StudentsInfo(
        #             name=row[0],
        #             address=row[2],
        #             rollno=row[3],
        #             gender=row[4],
        #             dob=dob,
        #             organization=organization,
        #             first_password=password,
        #             teacher=teacher,
        #         )
        #         student.parent = parent
        #         student.user = studentUser
        #         student.save()

        # messages.success(request, "Registration Completed")
        # return redirect("accounts:bulk_register")


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_teacher, login_url="accounts:forbidden")
@password_change_required
def downloadData(request):

    output = io.BytesIO()
    wb = xlsxwriter.Workbook(output)
    parentSheet = wb.add_worksheet("Parents Data")
    studentSheet = wb.add_worksheet("Students Data")

    # student sheet
    row_num = 0
    studentColumns = [
        "Name",
        "Roll No",
        "DOB",
        "Gender",
        "Address",
        "Organization",
        "Parent's Email",
        "Username",
        "Password",
    ]
    for col_num in range(len(studentColumns)):
        studentSheet.write(
            row_num, col_num, studentColumns[col_num]
        )  # at 0 row 0 column

    teacher = TeacherInCharge.objects.filter(user=request.user).first()
    rows = StudentsInfo.objects.filter(teacher=teacher).values_list(
        "name",
        "rollno",
        "dob",
        "gender",
        "address",
        "organization",
        "parent",
        "user",
        "first_password",
        "password_changed",
    )
    parentEmail = []
    for row in rows:
        row_num += 1
        for col_num in range(len(studentColumns)):
            if col_num == 0:
                studentSheet.write(
                    row_num, col_num, encryptionHelper.decrypt(row[col_num])
                )

            elif col_num == 2:
                studentSheet.write(row_num, col_num, row[col_num].strftime("%d/%b/%Y"))

            elif col_num == 5:
                organization = Organization.objects.get(pk=row[col_num])
                studentSheet.write(row_num, col_num, organization.name)

            elif col_num == 6:
                parent = ParentsInfo.objects.get(pk=row[col_num])
                email = encryptionHelper.decrypt(parent.email)
                if email not in parentEmail:
                    parentEmail.append(email)
                studentSheet.write(row_num, col_num, email)

            elif col_num == 7:
                user = User.objects.get(pk=row[col_num])
                studentSheet.write(row_num, col_num, user.username)

            elif col_num == 8:
                msg = row[col_num]
                if row[col_num + 1]:
                    msg = "Already Changed"
                studentSheet.write(row_num, col_num, msg)

            elif col_num == 9:
                continue

            else:
                studentSheet.write(row_num, col_num, row[col_num])

    # Sheet header, first row
    row_num = 0

    # parent sheet
    parentColumns = [
        "Name",
        "Email",
        "Age",
        "Gender",
        "Address",
        "City",
        "State",
        "Pincode",
        "Education",
        "Occupation",
        "Religion",
        "Family count",
        "Children count",
        "Family Type",
        "Username",
        "Password",
    ]
    for col_num in range(len(parentColumns)):
        parentSheet.write(row_num, col_num, parentColumns[col_num])  # at 0 row 0 column

    rows = ParentsInfo.objects.all().values_list(
        "name",
        "email",
        "age",
        "gender",
        "address",
        "city",
        "state",
        "pincode",
        "edu",
        "occupation",
        "religion",
        "no_of_family_members",
        "children_count",
        "type_of_family",
        "user",
        "first_password",
        "password_changed",
    )
    for row in rows:
        if encryptionHelper.decrypt(row[1]) in parentEmail:
            row_num += 1
            for col_num in range(len(parentColumns)):

                if col_num == 0 or col_num == 1:
                    parentSheet.write(
                        row_num, col_num, encryptionHelper.decrypt(row[col_num])
                    )

                elif col_num == 5:
                    city = City.objects.get(pk=row[col_num])
                    parentSheet.write(row_num, col_num, city.city)

                elif col_num == 6:
                    state = State.objects.get(pk=row[col_num])
                    parentSheet.write(row_num, col_num, state.state)

                elif col_num == 8:
                    education = Education.objects.get(pk=row[col_num])
                    parentSheet.write(row_num, col_num, education.education)

                elif col_num == 9:
                    occupation = Occupation.objects.get(pk=row[col_num])
                    parentSheet.write(row_num, col_num, occupation.occupation)

                elif col_num == 10:
                    religion = ReligiousBelief.objects.get(pk=row[col_num])
                    parentSheet.write(row_num, col_num, religion.religion)

                elif col_num == 13:
                    familyType = FamilyType.objects.get(pk=row[col_num])
                    parentSheet.write(row_num, col_num, familyType.family)

                elif col_num == 14:
                    user = User.objects.get(pk=row[col_num])
                    parentSheet.write(row_num, col_num, user.username)

                elif col_num == 15:
                    msg = row[col_num]
                    if row[col_num + 1]:
                        msg = "Already Changed"
                    parentSheet.write(row_num, col_num, msg)

                elif col_num == 16:
                    continue

                else:
                    parentSheet.write(row_num, col_num, row[col_num])

    wb.close()
    # construct response
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response[
        "Content-Disposition"
    ] = "attachment; filename=Parent and Student list.xlsx"
    return response


@login_required(login_url="accounts:loginlink")
@user_passes_test(
    lambda user: is_teacher(user),
    login_url="accounts:forbidden",
)
@password_change_required
def view_teacher_profile(request):
    if request.method == "GET":
        user = request.user
        if is_teacher(user):
            teacher = TeacherInCharge.objects.filter(user=user).first()

            teacher.fname = encryptionHelper.decrypt(teacher.fname)
            teacher.lname = encryptionHelper.decrypt(teacher.lname)

            if teacher.mname:
                teacher.mname = encryptionHelper.decrypt(teacher.mname)
                if teacher.mname == "":
                    teacher.mname = ""
            else:
                teacher.mname = ""

            if teacher.aadhar:
                teacher.aadhar = encryptionHelper.decrypt(teacher.aadhar)
                if teacher.aadhar == "":
                    teacher.aadhar = "-"
            else:
                teacher.aadhar = "-"

            if teacher.email:
                teacher.email = encryptionHelper.decrypt(teacher.email)
                if teacher.email == "":
                    teacher.email = "-"
            else:
                teacher.email = "-"

            if teacher.mobile_no:
                teacher.mobile_no = encryptionHelper.decrypt(teacher.mobile_no)
                if teacher.mobile_no == "":
                    teacher.mobile_no = "-"
            else:
                teacher.mobile_no = "-"

            teacher.dob = encryptionHelper.decrypt(teacher.dob)
            teacher.gender = encryptionHelper.decrypt(teacher.gender)

            return render(
                request,
                "teacher/view_teacher_profile.html",
                {"page_type": "view_teacher_profile", "teacher": teacher},
            )


@login_required(login_url="accounts:loginlink")
@user_passes_test(
    lambda user: is_teacher(user),
    login_url="accounts:forbidden",
)
@password_change_required
def edit_teacher_profile(request):
    teacher = TeacherInCharge.objects.filter(user=request.user).first()
    organization = teacher.organization
    if request.method == "GET":
        initial_dict = {
            "fname": encryptionHelper.decrypt(teacher.fname),
            "lname": encryptionHelper.decrypt(teacher.lname),
            "profile_pic": teacher.profile_pic,
            "dob": encryptionHelper.decrypt(teacher.dob),
            "gender": encryptionHelper.decrypt(teacher.gender),
        }

        if teacher.mname:
            initial_dict["mname"] = encryptionHelper.decrypt(teacher.mname)
        if teacher.aadhar:
            initial_dict["aadhar"] = encryptionHelper.decrypt(teacher.aadhar)
        if teacher.email:
            initial_dict["email"] = encryptionHelper.decrypt(teacher.email)
        if teacher.mobile_no:
            initial_dict["mobile_no"] = encryptionHelper.decrypt(teacher.mobile_no)

        form = TeachersInfoForm(request.POST or None, initial=initial_dict)
        return render(
            request,
            "teacher/update_teachers_info.html",
            {
                "form": form,
                "organization": organization,
            },
        )
    else:
        form = TeachersInfoForm(request.POST, request.FILES)

        if form.is_valid():
            teacher = TeacherInCharge.objects.filter(user=request.user).first()

            if teacher.mname:
                teacher.mname = encryptionHelper.encrypt(request.POST["mname"])
            if teacher.aadhar:
                teacher.aadhar = encryptionHelper.encrypt(request.POST["aadhar"])
            if teacher.email:
                teacher.email = encryptionHelper.encrypt(request.POST["email"])
            if teacher.mobile_no:
                teacher.mobile_no = encryptionHelper.encrypt(request.POST["mobile_no"])

            teacher.fname = encryptionHelper.encrypt(request.POST["fname"])
            teacher.lname = encryptionHelper.encrypt(request.POST["lname"])
            teacher.dob = encryptionHelper.encrypt(request.POST["dob"])
            teacher.gender = encryptionHelper.encrypt(request.POST["gender"])

            if request.FILES:
                if request.FILES["profile_pic"].size > 5 * 1024 * 1024:
                    form.add_error("profile_pic", "File size must be less than 5MB.")

                    return render(
                        request,
                        "teacher/update_teachers_info.html",
                        {
                            "form": form,
                            "organization": organization,
                        },
                    )
                else:
                    x = teacher.profile_pic.url.split("/account/media/accounts/")
                    if x[1] != "default.svg":
                        file = settings.MEDIA_ROOT + "/" + x[1]
                        os.remove(file)
                    teacher.profile_pic = request.FILES["profile_pic"]
            else:
                if "profile_pic-clear" in request.POST.keys():
                    x = teacher.profile_pic.url.split("/account/media/accounts/")
                    if x[1] != "default.svg":
                        file = settings.MEDIA_ROOT + "/" + x[1]
                        os.remove(file)
                    teacher.profile_pic = "/default.svg"

            teacher.save()
            return redirect("accounts:view_teacher_profile")
        else:

            print(request.POST)
            return render(
                request,
                "teacher/update_teachers_info.html",
                {"form": form, "organization": organization},
            )